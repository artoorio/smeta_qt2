import math
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


def export_with_fact_formula(df: pd.DataFrame, output_path: str = "output.xlsx"):
    df = df.copy()
    df["Количество(факт)"] = None

    wb = Workbook(); ws = wb.active
    headers = list(df.columns); ws.append(headers)

    # ---------- формулы «Количество(факт)» ----------
    last_work_row, last_work_qty = None, None
    cat_col = headers.index("Категория")
    qty_col = headers.index("Количество")
    fact_col = headers.index("Количество(факт)")
    fact_letter = get_column_letter(fact_col + 1)

    for idx, row in df.iterrows():
        excel_row = idx + 2
        category, qty = row["Категория"], row["Количество"]

        if category == "Работа":
            df.at[idx, "Количество(факт)"] = 0
            last_work_row, last_work_qty = excel_row, qty
        elif category == "Материалы" and last_work_row and last_work_qty:
            ratio = qty / last_work_qty if last_work_qty else 0
            df.at[idx, "Количество(факт)"] = f"={fact_letter}{last_work_row}*{round(ratio, 9)}"
        else:
            df.at[idx, "Количество(факт)"] = ""

    # ---------- запись DataFrame ----------
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)

    # ---------- стили ----------
    fill_work  = PatternFill("solid", fgColor="D6EFD6")
    fill_mat   = PatternFill("solid", fgColor="D6EAF8")
    fill_red   = PatternFill("solid", fgColor="F8D6D6")
    thin = Side(style="thin", color="000000")
    bottom_border = Border(bottom=thin)

    name_col_idx = headers.index("Наименование") + 1

    for excel_row in range(2, len(df) + 2):
        cat = ws.cell(row=excel_row, column=cat_col + 1).value
        qty = ws.cell(row=excel_row, column=qty_col + 1).value

        # цвет строки
        if cat == "Работа":
            target_fill = fill_work
        elif cat == "Материалы":
            target_fill = fill_red if qty is not None and qty < 0 else fill_mat
        elif cat == "Механизмы":
            target_fill = fill_work  # можно задать иной цвет при желании
        else:
            target_fill = None

        if target_fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=excel_row, column=col).fill = target_fill

        # перенос текста + высота строки
        cell_name = ws.cell(row=excel_row, column=name_col_idx)
        cell_name.alignment = Alignment(wrapText=True)
        text_len = len(str(cell_name.value))
        lines = max(1, math.ceil(text_len / 45))     # 45≈символов в 40-char колонке
        ws.row_dimensions[excel_row].height = lines * 15

        # граница под предыдущую строку, если текущая категория "Работа" или "Механизмы"
        if cat in ("Работа", "Механизмы"):
            for col in range(1, len(headers) + 1):
                ws.cell(row=excel_row - 1, column=col).border = bottom_border

    # ---------- ширина колонок ----------
    ws.column_dimensions[get_column_letter(name_col_idx)].width = 40
    ws.column_dimensions[get_column_letter(headers.index("Код расценки") + 1)].width = 15
    ws.column_dimensions[get_column_letter(cat_col + 1)].width = 10

    wb.save(output_path)
    print(f"Файл сохранён: {output_path}")