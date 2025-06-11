import openpyxl
import pandas as pd
import os
import logging
import warnings


# Настройка логирования
logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")
# Игнорирование отсутствие стиля
warnings.filterwarnings('ignore', category=UserWarning, message='Workbook contains no default style')

class Smeta:
    def __init__(self, file_path):
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Файл {file_path} не найден.")

        self.file_path = file_path
        try:
            self.wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            logging.error(f"Ошибка при загрузке файла: {e}")
            raise

        self.sheet = self.wb.active
        self.section_start_row = self._find_section_start()

        # Инициализируем DataFrame сразу
        self.data = pd.DataFrame(
            columns=[
                "Подраздел", "Номер позиции", "Код расценки", "Наименование",
                "Категория", "Единица измерения", "Количество", "Стоимость",
                "ФОТ", "ЭМ", "Материалы", "НР", "СП", "ОТм", "Вспомогательные ресурсы", "Оборудование"
            ]
        )

    def _find_section_start(self):
        """Ищет строку начала раздела (где указана 'Раздел 1.') в столбце A."""
        for row in self.sheet.iter_rows(min_row=1, max_col=1):
            cell_value = row[0].value
            if cell_value and isinstance(cell_value, str) and "Раздел 1." in cell_value:
                return row[0].row
        return None

    def _process_cost_details(self, row):
        """Обрабатывает строку с дополнительными данными."""
        cost_mapping = {
            "ОТ(ЗТ)": "ФОТ",
            "ЭМ": "ЭМ",
            "М": "Материалы",
            "ОТм(ЗТм)": "ОТм",
            "НР": "НР",
            "СП": "СП",
            "Вспомогательные ненормируемые материальные ресурсы": "Вспомогательные ресурсы",
        }
        details = {key: None for key in cost_mapping.values()}

        for key, target in cost_mapping.items():
            if isinstance(row[2], str) and row[2].startswith(key):
                details[target] = row[15] or 0
                break

        return details

    def _parse_data(self):
        """Парсит данные начиная с найденного раздела."""
        if not self.section_start_row:
            raise ValueError("Не удалось найти строку начала раздела.")

        rows = self.sheet.iter_rows(min_row=self.section_start_row, values_only=True)
        current_subsection = None
        parsing_position = False
        position_data = {}

        for row in rows:
            cell_a, cell_b, cell_c = row[0], row[1], row[2]

            # Определяем текущий подраздел
            if isinstance(cell_a, str) and not cell_b and not cell_c:
                current_subsection = cell_a
                continue

            # Если нашли новую позицию
            #if (isinstance(cell_a, (int, float)) or (isinstance(cell_a, str) and cell_a.isdigit())) and cell_b:
            if (isinstance(cell_a, (int, float)) or (isinstance(cell_a, str) and cell_a.isdigit())) or (isinstance(cell_a, str) and "\nО" in cell_a) and cell_b:
                # Если уже обрабатывалась предыдущая позиция, добавляем её в DataFrame
                if parsing_position:
                    self.data = pd.concat([self.data, pd.DataFrame([position_data])], ignore_index=True)

                # Определяем категорию
                category = self._get_category(cell_a, cell_b)

                # Инициализируем новую позицию
                parsing_position = True
                position_data = {
                    "Подраздел": current_subsection,
                    "Номер позиции": cell_a,
                    "Код расценки": cell_b,
                    "Наименование": cell_c,
                    "Категория": category,
                    "Единица измерения": row[7],
                    "Количество": row[10],
                    "Стоимость": None,
                    "ФОТ": None,
                    "ЭМ": None,
                    "Материалы": None,
                    "НР": None,
                    "СП": None,
                    "ОТм": None,
                    "Вспомогательные ресурсы": None,
                    "Оборудование": None,
                }
                continue

            # Если нашли строку "Всего по позиции"
            if parsing_position and cell_c == "Всего по позиции":
                position_data["Стоимость"] = row[15]

                # Записываем стоимость в нужный столбец в зависимости от категории
                if position_data["Категория"] == "Материалы":
                    position_data["Материалы"] = row[15]
                elif position_data["Категория"] == "Оборудование":
                    position_data["Оборудование"] = row[15]

                self.data = pd.concat([self.data, pd.DataFrame([position_data])], ignore_index=True)
                parsing_position = False
                continue

            # Обрабатываем дополнительные строки внутри текущей позиции
            if parsing_position:
                cost_details = self._process_cost_details(row)
                for key, value in cost_details.items():
                    if value is not None:
                        position_data[key] = (position_data[key] or 0) + value

    def _get_category(self, cell_a, cell_b):
        """Определяет категорию позиции."""
        category_mapping = {
            "ФСБЦ": "Материалы",
            "ТЦ": "Материалы",
            "ГЭСН": "Работа",
            "ФСЭМ": "Механизмы",
        }

        # Проверяем, содержит ли номер позиции схему "номер \n О"
        #print(cell_a)
        if isinstance(cell_a, str) and "\nО" in cell_a:
            return "Оборудование"

        # Проверяем код расценки
        for prefix, category in category_mapping.items():
            if isinstance(cell_b, str) and cell_b.startswith(prefix):
                return category

        return "Неизвестная категория"

    def export_to_excel(self, output_dir):
        """Экспортирует данные в Excel файл."""
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        output_file = os.path.join(output_dir, f"parsed_{os.path.basename(self.file_path)}")

        # Создаем новый Excel-файл с двумя листами
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Экспортируем основной DataFrame на первый лист
            self.data.to_excel(writer, index=False, sheet_name='Данные')

            # Создаем сводную таблицу
            pivot_table = pd.pivot_table(
                self.data,
                values='Стоимость',
                index='Подраздел',
                aggfunc='sum'
            ).reset_index()

            # Экспортируем сводную таблицу на второй лист
            #pivot_table
            smeta.data.groupby('Подраздел')["Стоимость"].sum().to_excel(writer, sheet_name='Сводная таблица')

        logging.info(f"Данные экспортированы в файл: {output_file}")


# Пример использования

directory = '/content/'

# Создаем пустой список для хранения названий файлов
xlsx_files = []

# Проходим по всем файлам в указанной директории
for filename in os.listdir(directory):
    # Проверяем, что файл имеет расширение .xlsx
    if filename.endswith('.xlsx'):
        # Добавляем название файла в список
        xlsx_files.append(filename)
for filepath in  xlsx_files:

    try:

        smeta = Smeta('/content/'+filepath)
        smeta._parse_data()  # Парсим данные
        #print(smeta.data)  # Выводим DataFrame

        # Экспорт данных в Excel
        output_dir = "output"  # Папка для сохранения результата
        smeta.export_to_excel(output_dir)
        print(smeta.data["Стоимость"].sum(), filepath)
    except Exception as e:
        logging.error(f"Произошла ошибка: {e}")