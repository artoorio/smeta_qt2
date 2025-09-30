#!pip install openpyxl

import os
import zlib


import openpyxl
import pandas as pd



def ch(a,b):
    ex = ['п.26.2']
    for c in ex:
        if c in b:
            return False

    for c in a:
        if c in b:
            return True
    return False

def str_to_float(a):
    if type(a)==str:
        a = float(a.replace(',', '.'))
    return a

class  Smeta(object):

    def __init__(self, file_name, okn):
        self.file_name = file_name
        if okn == 1:
            self.okn = 1.2
        else:
            self.okn = 1



    def check_indexes(self):
        excel_doc = openpyxl.open(filename=self.file_name, data_only=True)
        #excel_doc.sheetnames
        self.sheet = excel_doc[excel_doc.sheetnames[0]]
        self.kOT = 0
        self.kEM = 0
        self.kM = 0
        self.ws_begining = 0
        self.ws_end = 0
        kEM1 = 0
        kM1 = 0
        for i in range(1,self.sheet.max_row):

            if type(self.sheet[f'C{i}'].value)==str:
                if 'Наименование работ и затрат' in self.sheet[f'C{i}'].value:
                    #print(f'Наименование столбцов таблицы строка {i}')
                    self.ws_begining = i #строка с оглавлением таблицы

            if type(self.sheet[f'C{i}'].value)==str:
                if 'ВСЕГО по смете' in self.sheet[f'C{i}'].value or 'ВСЕГО по смете' in self.sheet[f'C{i}'].value:

                    self.ws_end = i #строка с ВСЕГО по смете
                    break


        for i in range(self.ws_begining,self.ws_end):

            if type(self.sheet[f'C{i}'].value)==str:
                if 'ОТ'==self.sheet[f'C{i}'].value:
                    self.kOT = self.sheet[f'M{i}'].value
                    #print(f'коэффициент на оплату труда {self.kOT}')
                    break

        for i in range(self.ws_end,self.ws_begining,-1):

            if type(self.sheet[f'C{i}'].value)==str:

                if 'М'==self.sheet[f'C{i}'].value:
                    self.kM = self.sheet[f'M{i}'].value
                    #print(f'коэффициент на материалы {kM}')
                    break

        for i in range(self.ws_end,self.ws_begining,-1):

            if type(self.sheet[f'C{i}'].value)==str:
                if 'ЭМ'==self.sheet[f'C{i}'].value:
                    self.kEM = self.sheet[f'M{i}'].value
                    #print(f'коэффициент на механизмы {kEM}')
                    break

        for i in range(self.ws_end,self.ws_begining,-1):

            if type(self.sheet[f'C{i}'].value)==str:

                if 'материалы' in self.sheet[f'C{i}'].value:
                    kM1 = self.sheet[f'M{i}'].value
                    #print(f'коэффициент на материалы {kM1}')
                    break

        for i in range(self.ws_end,self.ws_begining,-1):

            if type(self.sheet[f'C{i}'].value)==str:
                if 'эксплуатация машин и механизмов' in self.sheet[f'C{i}'].value:
                    kEM1 = self.sheet[f'M{i}'].value
                    #print(f'коэффициент на механизмы {kEM1}')
                    break

        if not self.kEM: self.kEM = kEM1
        if not self.kM: self.kM = kM1


        if type(self.kOT)==str:
            self.kOT = float(self.kOT.replace(',', '.'))
        if type(self.kEM)==str:
            self.kEM = float(self.kEM.replace(',', '.'))
        if type(self.kM)==str:
            self.kM = float(self.kM.replace(',', '.'))

        #self.kOT, self.kEM, self.kM = 32.46, 12.13, 8.52
        #self.kOT, self.kEM, self.kM = 27.17, 10.33, 7.53
        print(self.kOT, self.kEM, self.kM)




    def check_xls(self):

        self.docs = openpyxl.Workbook()
        self.docs.create_sheet(title = 'Лист1', index = 0)
        self.docs.create_sheet(title = 'Лист2', index = 1)
        sheetnames =self.docs.sheetnames
        ws2 = self.docs[sheetnames[0]]
        ws3 = self.docs[sheetnames[1]]

        self.sum = 0
        self.sum_mat=0
        self.sum_rab=0
        self.sum_per=0
        self.sum_nrsp=0
        self.sum_comment = 0
        self.sheet1 = 0
        self.sheet2 = 0
        nomer = 1
        nomer2 = 1
        flag = 0

        ws3[f'A{nomer2}'] = 'Номер'
        ws3[f'B{nomer2}'] = 'Наименование подраздела'
        ws3[f'C{nomer2}'] = 'стоимость'
        ws3[f'D{nomer2}'] = 'стоимость материалов по смете'
        ws3.column_dimensions['B'].width=60
        ws3.column_dimensions['C'].width=20
        ws3.column_dimensions['D'].width=60

        ws2[f'A{nomer}'] = 'Номер'
        ws2[f'B{nomer}'] = 'Наименование позиции'   #наименование позиции
        ws2[f'C{nomer}'] = 'Единица измерения'   #единица измерения
        ws2[f'D{nomer}'] = 'Количество'   #количество
        ws2[f'E{nomer}'] = 'Цена без НДС'   #цена
        ws2[f'F{nomer}'] = 'Цена с НДС'
        ws2[f'G{nomer}'] = 'Цена за единицу'
        ws2[f'H{nomer}'] = 'Категория'
        ws2[f'I{nomer}'] = 'Обоснование'
        ws2[f'K{nomer}'] = 'Работы(на ед.)'
        ws2[f'J{nomer}'] = 'Материалы(на ед.)'
        ws2[f'L{nomer}'] = 'Подраздел'
        ws2.column_dimensions['B'].width=60
        ws2.column_dimensions['C'].width=20
        ws2.column_dimensions['D'].width=12
        ws2.column_dimensions['E'].width=14
        ws2.column_dimensions['F'].width=14
        ws2.column_dimensions['G'].width=14
        ws2.column_dimensions['H'].width=12

        nomer = 2
        #nomer2 = 2
        poz = ['ТЕР', 'ТССЦ', 'ТЦ', 'ФЕР', 'ФССЦ','ТСЭМ']
        rab =  ['ТЕР','ФЕР']
        kac = ['ТЦ','КА','КА','Кон']  # текущие цены

        for i in range(self.ws_begining,self.ws_end): #проходим смету начиная с оглавления таблицы

            if type(self.sheet[f'A{i}'].value)==str:
                if 'Раздел 1' in self.sheet[f'A{i}'].value:
                    flag = 1
            if type(self.sheet[f'A{i}'].value)==str:
                if (self.sheet[f'A{i}'].value.isdigit() == False) and (flag==1) and (not ch(['Н','Уд'],self.sheet[f'A{i}'].value)):
                    nomer2 += 1
                    ws3[f'A{nomer2}'] = nomer2-1
                    ws3[f'B{nomer2}'] = self.sheet[f'A{i}'].value
                    #ws2[f'L{nomer}'] = ws3[f'B{nomer2}']
                    ws3[f'C{nomer2}'].value = 0
                    ws3[f'C{nomer2}'].number_format = '#,##0.00'
                    ws3[f'D{nomer2}'].value = 0
                    ws3[f'D{nomer2}'].number_format = '#,##0.00'


            if type(self.sheet[f'B{i}'].value)==str:


                if ch(['ТЕР', 'ТССЦ', 'ТЦ', 'ФЕР', 'ФССЦ', 'Кон', 'КА','ТСЭ','Цпг','ФСЭМ', 'ТСЭМ'], self.sheet[f'B{i}'].value): #если встретили в столбце C
                    ws2[f'L{nomer}'] = ws3[f'B{nomer2}'].value
                    if ch(['ТЕР','ФЕР'], self.sheet[f'B{i}'].value): #ТЕР, ФЕР
                        if self.sheet[f'I{i}'].value==0:
                                continue
                        ws2[f'A{nomer}'] = nomer-1
                        ws2[f'B{nomer}'] = self.sheet[f'C{i}'].value   #наименование позиции
                        ws2[f'C{nomer}'] = self.sheet[f'F{i}'].value   #единица измерения
                        ws2[f'D{nomer}'] = str_to_float(self.sheet[f'I{i}'].value)   #количество
                        ot = 0
                        em = 0
                        vtOT, mat, nr, sp = 0, 0, 0, 0,
                        #print('haha')
                        for k in range(i, self.ws_end):
                            if type(self.sheet[f'C{k}'].value)==str:
                                if 'Всего по позиции' in self.sheet[f'C{k}'].value:
                                    break
                                else:
                                    if 'ОТ' == self.sheet[f'C{k}'].value:
                                        ot = self.sheet[f'L{k}'].value
                                    if 'ЭМ' == self.sheet[f'C{k}'].value:
                                        em = self.sheet[f'L{k}'].value
                                    if 'в т.ч. ОТм' == self.sheet[f'C{k}'].value:
                                        vtOT = self.sheet[f'L{k}'].value
                                    if 'М' == self.sheet[f'C{k}'].value:
                                        mat = self.sheet[f'L{k}'].value
                                    if 'НР ' in self.sheet[f'C{k}'].value:
                                        nr = self.sheet[f'L{k}'].value
                                        if self.sheet[f'I{k}'].value==0:
                                            nr = 0
                                    if 'СП ' in self.sheet[f'C{k}'].value:
                                        sp = self.sheet[f'L{k}'].value
                                        if self.sheet[f'I{k}'].value==0:
                                            sp = 0

                        #print(type(ot), type(em),type(mat), type(nr), type(sp))
                        '''if str(sp).isdigit()==False:
                            sp = 0
                            print(nomer, 'в позиции СП неверно')
                        if str(nr).isdigit()==False:
                            nr = 0
                            print(nomer, 'в позиции НР неверно')'''
                        if not nr: nr = 0
                        if not sp: sp = 0
                        ws2[f'E{nomer}'] = ot*self.kOT+em*self.kEM+mat*self.kM+nr*self.kOT+sp*self.kOT
                        #ws2[f'F{nomer}'] = (ot*self.kOT+em*self.kEM+nr*self.kOT+sp*self.kOT)*1.2/self.okn+mat*self.kM*1.2
                        ws2[f'F{nomer}'] = (ot*self.kOT+vtOT*self.kOT+nr*self.kOT+sp*self.kOT)*1.2/self.okn+mat*self.kM*1.2+(em*self.kEM-vtOT*self.kOT)*1.2
                        if ws2[f'D{nomer}'].value==0 or not ws2[f'D{nomer}'].value:
                            ws2[f'G{nomer}'] = 0
                        else:

                            ws2[f'G{nomer}'] = ws2[f'F{nomer}'].value/ws2[f'D{nomer}'].value
                        ws2[f'H{nomer}'] = 'работа'
                        ws2[f'E{nomer}'].number_format = '#,##0.00'
                        ws2[f'F{nomer}'].number_format = '#,##0.00'
                        ws2[f'G{nomer}'].number_format = '#,##0.00'
                        ws2[f'I{nomer}'] = self.sheet[f'B{i}'].value
                        self.sum_mat += mat*self.kM*1.2
                        #self.sum_rab += (ot*self.kOT+em*self.kEM+nr*self.kOT+sp*self.kOT)*1.2/self.okn
                        self.sum_rab += (ot*self.kOT+vtOT*self.kOT+nr*self.kOT+sp*self.kOT)*1.2/self.okn+(em*self.kEM-vtOT*self.kOT)*1.2
                        self.sum_nrsp += (nr*self.kOT+sp*self.kOT)*1.2*self.okn
                        #self.sum += (ot*self.kOT+em*self.kEM+nr*self.kOT+sp*self.kOT)*1.2/self.okn+mat*self.kM*1.2
                        self.sum +=  (ot*self.kOT+vtOT*self.kOT+nr*self.kOT+sp*self.kOT)*1.2/self.okn+mat*self.kM*1.2+(em*self.kEM-vtOT*self.kOT)*1.2
                        #print(ws2[f'F{nomer}'].value, ws3[f'C{nomer2}'].value)
                        #if not ws3[f'C{nomer2}'].value: ws3[f'C{nomer2}'].value=0
                        ws3[f'C{nomer2}'].value += ws2[f'F{nomer}'].value

                        #print(ot,em,mat,nr,sp)
                        ws2[f'K{nomer}'] = (ot*self.kOT+vtOT*self.kOT+nr*self.kOT+sp*self.kOT)*1.2/self.okn+(em*self.kEM-vtOT*self.kOT)*1.2
                        ws2[f'J{nomer}'] = mat*self.kM*1.2

                        ws3[f'D{nomer2}'].value += ws2[f'J{nomer}'].value

                        if ws2[f'D{nomer}'].value==0 or not ws2[f'D{nomer}'].value:
                            ws2[f'K{nomer}'] = 0
                            ws2[f'J{nomer}'] = 0
                        else:
                            ws2[f'K{nomer}'] =  ws2[f'K{nomer}'].value/ws2[f'D{nomer}'].value
                            ws2[f'J{nomer}'] =  ws2[f'J{nomer}'].value/ws2[f'D{nomer}'].value

                    if ch(['Цпг','ТСЭ','ФСЭМ','ТСЭМ'],self.sheet[f'B{i}'].value):
                        if self.sheet[f'I{i}'].value==0:
                                continue
                        ws2[f'A{nomer}'] = nomer-1
                        ws2[f'B{nomer}'] = self.sheet[f'C{i}'].value   #наименование позиции
                        ws2[f'C{nomer}'] = self.sheet[f'F{i}'].value   #единица измерения
                        ws2[f'D{nomer}'] = self.sheet[f'I{i}'].value   #количество
                        ws2[f'E{nomer}'] = self.sheet[f'L{i}'].value*self.kEM   #цена
                        ws2[f'F{nomer}'] = self.sheet[f'L{i}'].value*self.kEM*1.2/self.okn

                        if ws2[f'D{nomer}'].value==0 or not ws2[f'D{nomer}'].value:
                            ws2[f'G{nomer}'] = 0
                        else:
                            ws2[f'F{nomer}'].value = str_to_float(ws2[f'F{nomer}'].value)
                            ws2[f'D{nomer}'].value = str_to_float(ws2[f'D{nomer}'].value)
                            ws2[f'G{nomer}'] = ws2[f'F{nomer}'].value/ws2[f'D{nomer}'].value
                        ws2[f'H{nomer}'] = 'перевозки'
                        ws2[f'I{nomer}'] = self.sheet[f'B{i}'].value
                        ws2[f'E{nomer}'].number_format = '#,##0.00'
                        ws2[f'F{nomer}'].number_format = '#,##0.00'
                        ws2[f'G{nomer}'].number_format = '#,##0.00'
                        self.sum_per += self.sheet[f'L{i}'].value*self.kEM*1.2/self.okn
                        self.sum += self.sheet[f'L{i}'].value*self.kEM*1.2/self.okn
                        ws3[f'C{nomer2}'].value += ws2[f'F{nomer}'].value
                        ws2[f'K{nomer}'] = self.sheet[f'L{i}'].value*self.kEM*1.2/self.okn
                        ws2[f'K{nomer}'] = ws2[f'G{nomer}'].value
                    elif ch(['ТССЦ','ФССЦ'],self.sheet[f'B{i}'].value):
                        if self.sheet[f'I{i}'].value==0:
                            continue
                        ws2[f'A{nomer}'] = nomer-1
                        ws2[f'B{nomer}'] = self.sheet[f'C{i}'].value   #наименование позиции
                        ws2[f'C{nomer}'] = self.sheet[f'F{i}'].value   #единица измерения
                        ws2[f'D{nomer}'] = self.sheet[f'I{i}'].value   #количество

                        if not self.sheet[f'L{i}'].value: self.sheet[f'L{i}'].value = 0
                        ws2[f'E{nomer}'] = self.sheet[f'L{i}'].value*self.kM   #цена

                        ws2[f'F{nomer}'] = self.sheet[f'L{i}'].value*self.kM*1.2
                        if ws2[f'D{nomer}'].value==0 or not ws2[f'D{nomer}'].value:
                            ws2[f'G{nomer}'] = 0
                        else:
                            ws2[f'F{nomer}'].value = str_to_float(ws2[f'F{nomer}'].value)
                            ws2[f'D{nomer}'].value = str_to_float(ws2[f'D{nomer}'].value)
                            ws2[f'G{nomer}'] = ws2[f'F{nomer}'].value/ws2[f'D{nomer}'].value
                        ws2[f'H{nomer}'] = 'материал'
                        ws2[f'I{nomer}'] = self.sheet[f'B{i}'].value
                        ws2[f'E{nomer}'].number_format = '#,##0.00'
                        ws2[f'F{nomer}'].number_format = '#,##0.00'
                        ws2[f'G{nomer}'].number_format = '#,##0.00'
                        self.sum += self.sheet[f'L{i}'].value*self.kM*1.2
                        self.sum_mat += self.sheet[f'L{i}'].value*self.kM*1.2
                        ws3[f'C{nomer2}'].value += ws2[f'F{nomer}'].value
                        ws3[f'D{nomer2}'].value += ws2[f'F{nomer}'].value
                        if ws2[f'D{nomer}'].value==0 or not ws2[f'D{nomer}'].value:
                            ws2[f'K{nomer}'] = 0
                            ws2[f'J{nomer}'] = 0
                        else:
                            #ws2[f'K{nomer}'] =  ws2[f'K{nomer}'].value/ws2[f'D{nomer}'].value
                            if ws2[f'J{nomer}'].value:
                                ws2[f'J{nomer}'] =  ws2[f'J{nomer}'].value/ws2[f'D{nomer}'].value
                        ws2[f'J{nomer}'] = ws2[f'G{nomer}'].value
                        ws2[f'K{nomer}'] = 0
                        #print(nomer,ws2[f'D{nomer}'].value )
                        if ws2[f'D{nomer}'].value<0: #если минусовой материал, то внутри работы вычитает стоимость материла
                            f = nomer

                            while  ws2[f'K{f}'].value==0:
                                f -= 1
                            #print(ws2[f'J{f}'].value,ws2[f'J{nomer}'].value,ws2[f'D{nomer}'].value,ws2[f'D{f}'].value)
                            if ws2[f'J{f}'].value==None:
                                ws2[f'J{f}'].value=0


                            ws2[f'J{f}'] = ws2[f'J{f}'].value +(ws2[f'J{nomer}'].value*ws2[f'D{nomer}'].value)/ ws2[f'D{f}'].value
                            ws2[f'J{nomer}'].value = 0
                    elif ch(['ТЦ','КА','КА','Кон'],self.sheet[f'B{i}'].value):
                        if self.sheet[f'I{i}'].value==0:
                            continue
                        ws2[f'A{nomer}'] = nomer-1
                        ws2[f'B{nomer}'] = self.sheet[f'C{i}'].value   #наименование позиции
                        ws2[f'C{nomer}'] = self.sheet[f'F{i}'].value   #единица измерения
                        ws2[f'D{nomer}'] = self.sheet[f'I{i}'].value   #количество
                        koef = 0
                        if self.sheet[f'M{i}']:
                            #print(self.sheet[f'M{i}'].value)
                            #if type(self.sheet[f'M{i}'].value)==str:
                                #print('popal')
                            if self.sheet[f'M{i}'].value==5.34:
                                    koef = self.kM
                                    self.kM = 5.34

                        if self.sheet[f'M{i}']:
                            self.sheet[f'M{i}'].value = str_to_float(self.sheet[f'M{i}'].value)
                            ws2[f'E{nomer}'] = self.sheet[f'L{i}'].value*self.sheet[f'M{i}'].value   #цена. заменил с self.kM
                            #print(ws2[f'E{nomer}'].value, self.sheet[f'L{i}'].value, self.sheet[f'M{i}'].value)

                            ws2[f'F{nomer}'] = ws2[f'E{nomer}'].value*1.2
                        if ws2[f'D{nomer}'].value==0 or not ws2[f'D{nomer}'].value:
                            ws2[f'G{nomer}'] = 0
                        else:
                            ws2[f'F{nomer}'].value = str_to_float(ws2[f'F{nomer}'].value)
                            ws2[f'D{nomer}'].value = str_to_float(ws2[f'D{nomer}'].value)
                            ws2[f'G{nomer}'] = ws2[f'F{nomer}'].value/ws2[f'D{nomer}'].value
                        ws2[f'H{nomer}'] = 'материал'
                        ws2[f'I{nomer}'] = self.sheet[f'B{i}'].value
                        ws2[f'E{nomer}'].number_format = '#,##0.00'
                        ws2[f'F{nomer}'].number_format = '#,##0.00'
                        ws2[f'G{nomer}'].number_format = '#,##0.00'

                        #self.sum += self.sheet[f'L{i}'].value*self.kM*1.2
                        #self.sum_mat += self.sheet[f'L{i}'].value*self.kM*1.2
                        self.sum += self.sheet[f'L{i}'].value*self.sheet[f'M{i}'].value*1.2
                        self.sum_mat += self.sheet[f'L{i}'].value*self.sheet[f'M{i}'].value*1.2
                        ws3[f'C{nomer2}'].value += ws2[f'F{nomer}'].value
                        ws3[f'D{nomer2}'].value += ws2[f'F{nomer}'].value
                        if ws2[f'D{nomer}'].value==0 or not ws2[f'D{nomer}'].value:
                            ws2[f'K{nomer}'] = 0
                            ws2[f'J{nomer}'] = 0
                        else:
                            #ws2[f'K{nomer}'] =  ws2[f'K{nomer}'].value/ws2[f'D{nomer}'].value
                            if ws2[f'J{nomer}'].value:
                                ws2[f'J{nomer}'] =  ws2[f'J{nomer}'].value/ws2[f'D{nomer}'].value
                        ws2[f'J{nomer}'] = ws2[f'G{nomer}'].value
                        ws2[f'K{nomer}'] = 0
                        if ws2[f'D{nomer}'].value<0:
                            f = nomer
                            while  ws2[f'K{f}']==0:
                                f -= 1
                            ws2[f'J{nomer}'] = ws2[f'J{nomer}'].value*ws2[f'D{nomer}'].value/ ws2[f'D{f}'].value
                        if koef>0: self.kM = koef





                    nomer += 1
        self.docs.save('_'+self.file_name+'_output.xlsx')
        self.docs.save('_Тест.xlsx')
        print('{0:,}'.format((round(self.sum,2))).replace(',', ' '))
        print('Стоимость материалов {0:,}'.format((round(self.sum_mat,2))).replace(',', ' '))
        print('Стоимость работ {0:,}'.format((round(self.sum_rab,2))).replace(',', ' ')+'(в том числе НР и СП:{0:,})'.format((round(self.sum_nrsp,2))).replace(',', ' '))
        print('Стоимость перевозок {0:,}'.format((round(self.sum_per,2))).replace(',', ' '))
        print('')


import os
import pandas as pd

def process_all_smet_files(path="."):
    pd.options.display.float_format = '{0:.2f}'.format
    file_types = {'.xlsx', '.xls'}

    dir_list = os.listdir(path)
    list_xlsx = []
    total_sum = 0
    k = 0

    # Собираем список файлов-смет
    for f in dir_list:
        filename, file_extension = os.path.splitext(f)
        if file_extension in file_types and filename[0] != '_':
            list_xlsx.append((f, filename))

    # Пробуем загрузить файл материалов
    materials_data = None
    materials_file = '_материалы.xlsx'
    if materials_file in dir_list:
        try:
            materials_data = pd.read_excel(materials_file)
            materials_data = materials_data[materials_data['Факт за ед'] > 0]
            materials_data.rename(columns={'Наименование работ и затрат': 'Наименование позиции'}, inplace=True)
            materials_data = materials_data.drop_duplicates(subset=['Наименование позиции'])
            print(f"✅ Загружен файл материалов: {materials_file}")
        except Exception as e:
            print(f"⚠️ Ошибка при загрузке файла {materials_file}: {e}")
            materials_data = None
    else:
        print(f"⚠️ Файл {materials_file} не найден. Фактические стоимости материалов считаться не будут.")

    list_df = []
    list_df2 = []

    for f, name in list_xlsx:
        print(f"\n▶️ Обработка: {f}")

        smet = Smeta(f, 0)
        smet.check_indexes()
        smet.check_xls()

        df1 = pd.read_excel('_Тест.xlsx', sheet_name='Лист1')
        df2 = pd.read_excel('_Тест.xlsx', sheet_name='Лист2')
        df1['Название объекта'] = f
        df2['Название объекта'] = f

        if materials_data is not None:
            df1['Факт за ед'] = df1['Наименование позиции'].map(
                materials_data.set_index('Наименование позиции')['Факт за ед'].to_dict()
            )
            df1['Смета за ед'] = df1['Наименование позиции'].map(
                materials_data.set_index('Наименование позиции')['Смета за ед'].to_dict()
            )
            df1['Смета материалы'] = df1['Материалы(на ед.)'] * df1['Количество']
            df1['Факт материалы'] = df1['Факт за ед'] * df1['Количество']

            smet_mat = df1['Смета материалы'].sum()
            fakt_mat = df1['Факт материалы'].sum()
            anal_mat = df1['Смета материалы'][df1['Факт материалы'] > 0].sum()

            print('Проанализировано материалов: {0:,}'.format(round(anal_mat, 2)).replace(',', ' '),
                  f" ({round(anal_mat * 100 / smet_mat, 2)}%)")
            print('Стоимость покупки материалов: {0:,}'.format(round(fakt_mat, 2)).replace(',', ' '),
                  f" ({round(fakt_mat * 100 / smet.sum, 2)}%)")
        else:
            print("⏩ Пропущен анализ фактической стоимости (нет файла материалов)")

        list_df.append(df1)
        list_df2.append(df2)
        total_sum += smet.sum
        k += 1

    # Объединение и сохранение результатов
    if list_df:
        pd.concat(list_df, ignore_index=True).to_excel("_output.xlsx", index=False)
    if list_df2:
        pd.concat(list_df2, ignore_index=True).to_excel("_output2.xlsx", index=False)

    print('\n🧾 Общая сумма по сметам: {0:,}'.format(round(total_sum, 2)).replace(',', ' '))


process_all_smet_files(path=".")