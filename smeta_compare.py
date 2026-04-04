import os
from difflib import SequenceMatcher
from typing import List, Union, Optional
import re
import numpy as np
import pandas as pd
from difflib import SequenceMatcher
import numpy as np
import pandas as pd
from data_processing import process_smeta
from export_formatting import apply_named_column_widths, apply_readable_sheet_layout, apply_section_row_grouping, cap_row_heights, autofit_columns_by_name
import math
from openpyxl.utils import get_column_letter     # ширина колонок
from openpyxl.styles import Alignment            # перенос + высота
class SmetaComparator:
    """Сравнение двух смет‑таблиц и генерация трёх типов отчётов.

    * **Стандартный** – обе сметы + разница (HTML/Excel)
    * **Объединённый** – колонки двух файлов рядом (HTML/Excel)
    * **Для заказчика** – упрощённый отчёт (HTML/Excel)

    При наличии колонки *subsection_column* в исходных данных вставляются
    строки‑заголовки вида ``-- <Подраздел> --``.
    """

    # ------------------------------------------------------------------
    # Init
    # ------------------------------------------------------------------
    def __init__(
        self,
        df1: pd.DataFrame,
        df2: pd.DataFrame,
        *,
        file1_name: str = "Файл 1",
        file2_name: str = "Файл 2",
        compare_column: str = "Наименование",
        value_column: Union[str, List[str]] = "Стоимость",
        extra_column: Optional[List[str]] = None,
        subsection_column: Optional[str] = "Подраздел",
    ) -> None:
        self.file1_name = os.path.splitext(os.path.basename(file1_name))[0]
        self.file2_name = os.path.splitext(os.path.basename(file2_name))[0]
        self.compare_column = compare_column
        self.value_column = [value_column] if isinstance(value_column, str) else value_column
        self.extra_column = extra_column or []
        self.subsection_column = subsection_column

        self.df1 = df1.copy().reset_index(drop=True)
        self.df2 = df2.copy().reset_index(drop=True)

        # clean numeric columns
        for df in (self.df1, self.df2):
            if self.compare_column not in df.columns:
                df[self.compare_column] = ""
            for col in self.value_column:
                if col not in df.columns:
                    df[col] = None
                df[col] = (
                    df[col].astype(str)
                    .str.replace(r"[^\d,.-]", "", regex=True)
                    .str.replace(",", ".", regex=False)
                )
                df[col] = pd.to_numeric(df[col], errors="coerce")

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _empty_with_zero_from_row(self, row, value_column):
        """
        Создает копию строки row с 0 в value_column.
        Остальные колонки остаются без изменений.
        """
        new_row = row.copy()
        if value_column in new_row:
            new_row[value_column] = 0
        return new_row

    def _align_subsection_safe(self):
        """
        Выравнивание df1 и df2 по подразделам с поддержкой дубликатов.
        Отсутствующие строки получают value_column=0.
        """
        col_sub = self.subsection_column
        val_col = self.value_column

        # работаем на копиях и нормализуем пропуски подразделов
        df1 = self.df1.copy()
        df2 = self.df2.copy()
        sentinel = "∅"
        df1[col_sub] = df1[col_sub].fillna(sentinel)
        df2[col_sub] = df2[col_sub].fillna(sentinel)

        # порядок подразделов: сначала df1, затем недостающие из df2
        order = list(dict.fromkeys(df1[col_sub].tolist()))
        for s in df2[col_sub].tolist():
            if s not in order:
                order.append(s)

        out1_blocks, out2_blocks = [], []

        for s in order:
            sub1 = df1[df1[col_sub] == s].reset_index(drop=True)
            sub2 = df2[df2[col_sub] == s].reset_index(drop=True)

            # SequenceMatcher по колонке сравнения
            list1 = sub1[self.compare_column].fillna("").astype(str).tolist()
            list2 = sub2[self.compare_column].fillna("").astype(str).tolist()
            sm = SequenceMatcher(None, list1, list2)

            r1, r2 = [], []

            for tag, i1, i2, j1, j2 in sm.get_opcodes():
                if tag == "equal":
                    r1.extend(sub1.iloc[i1:i2].to_dict("records"))
                    r2.extend(sub2.iloc[j1:j2].to_dict("records"))
                elif tag in ("replace", "delete", "insert"):
                    # строки из df1
                    for i in range(i1, i2):
                        r1.append(sub1.iloc[i].to_dict())
                        if j1 < j2:
                            # подставляем строку df2 с 0 в value_column
                            dummy = self._empty_with_zero_from_row(sub2.iloc[j1].to_dict(), val_col)
                            r2.append(dummy)
                            j1 += 1
                        else:
                            dummy = self._empty_with_zero_from_row(sub2.iloc[0].to_dict(),
                                                                   val_col) if not sub2.empty else {col: None for col in
                                                                                                    sub2.columns}
                            r2.append(dummy)

                    # строки из df2, которые остались
                    for j in range(j1, j2):
                        dummy = self._empty_with_zero_from_row(sub1.iloc[0].to_dict(), val_col) if not sub1.empty else {
                            col: None for col in sub1.columns}
                        r1.append(dummy)
                        r2.append(sub2.iloc[j].to_dict())

            # создаем блоки DataFrame с dtype=object
            a1 = pd.DataFrame(r1, dtype=object).reindex(columns=sub1.columns)
            a2 = pd.DataFrame(r2, dtype=object).reindex(columns=sub2.columns)

            out1_blocks.append(a1)
            out2_blocks.append(a2)

        # объединяем все подразделы
        out1 = pd.concat(out1_blocks, ignore_index=True) if out1_blocks else df1.iloc[0:0].copy()
        out2 = pd.concat(out2_blocks, ignore_index=True) if out2_blocks else df2.iloc[0:0].copy()

        # восстановление NaN вместо sentinel
        out1[col_sub] = out1[col_sub].replace(sentinel, np.nan)
        out2[col_sub] = out2[col_sub].replace(sentinel, np.nan)

        return out1, out2

    def _align_pair4(self, df1: pd.DataFrame, df2: pd.DataFrame):
        """
        Последовательное выравнивание двух подтаблиц по self.compare_column.
        Элементы из df2 вставляются на позиции первых встреченных совпадений в df1.
        """
        try:
            # Проверка входных данных
            if not isinstance(df1, pd.DataFrame) or not isinstance(df2, pd.DataFrame):
                raise TypeError("Both arguments must be pandas DataFrames")

            if self.compare_column not in df1.columns and not df1.empty:
                raise KeyError(f"Column '{self.compare_column}' not found in df1")

            if self.compare_column not in df2.columns and not df2.empty:
                raise KeyError(f"Column '{self.compare_column}' not found in df2")

            # Обработка граничных случаев
            if df1.empty and df2.empty:
                return df1.copy(), df2.copy()

            if df1.empty:
                # Исправлено: используем правильные колонки
                empty_rows = []
                for _ in range(len(df2)):
                    empty_rows.append(self._empty_like(df1))
                empty_df1 = pd.DataFrame(empty_rows)
                if not df1.columns.empty:
                    empty_df1 = empty_df1.reindex(columns=df1.columns, fill_value=np.nan)
                return empty_df1, df2.copy()

            if df2.empty:
                empty_rows = []
                for _ in range(len(df1)):
                    empty_rows.append(self._empty_like(df2))
                empty_df2 = pd.DataFrame(empty_rows)
                if not df2.columns.empty:
                    empty_df2 = empty_df2.reindex(columns=df2.columns, fill_value=np.nan)
                return df1.copy(), empty_df2

            # Ограничение размера для безопасности
            MAX_OPERATIONS = 100000
            if len(df1) * len(df2) > MAX_OPERATIONS:
                print(f"WARNING: Large dataset detected ({len(df1)} x {len(df2)}), processing may be slow")

            # Создаем списки для результата
            r1, r2 = [], []

            # Создаем очередь элементов df2 для сопоставления
            df2_queue = df2.copy().reset_index(drop=True)
            df2_used = [False] * len(df2_queue)

            # Проходим по df1 и ищем совпадения в df2_queue
            for idx, row1 in df1.iterrows():
                val1 = str(row1[self.compare_column]).strip()

                # Ищем первый неиспользованный элемент в df2_queue
                match_found = False
                for j in range(len(df2_queue)):
                    if df2_used[j]:  # уже использован
                        continue

                    val2 = str(df2_queue.iloc[j][self.compare_column]).strip()
                    if val1 == val2:
                        # Найдено совпадение
                        r1.append(row1.to_dict())
                        r2.append(df2_queue.iloc[j].to_dict())
                        df2_used[j] = True
                        match_found = True
                        break

                if not match_found:
                    # Совпадение не найдено - добавляем пустую строку для df2
                    r1.append(row1.to_dict())
                    r2.append(self._empty_like(df2))

            # Добавляем оставшиеся неиспользованные элементы df2
            for j in range(len(df2_queue)):
                if not df2_used[j]:
                    r1.append(self._empty_like(df1))
                    r2.append(df2_queue.iloc[j].to_dict())

            # Создаем результирующие DataFrame
            a1 = pd.DataFrame(r1)
            a2 = pd.DataFrame(r2)

            # Безопасное переиндексирование
            if not df1.columns.empty:
                a1 = a1.reindex(columns=df1.columns, fill_value=np.nan)
            if not df2.columns.empty:
                a2 = a2.reindex(columns=df2.columns, fill_value=np.nan)

            return a1.reset_index(drop=True), a2.reset_index(drop=True)

        except Exception as e:
            print(f"Error in _align_pair: {e}")
            import traceback
            traceback.print_exc()
            # Возвращаем безопасные пустые результаты
            return df1.iloc[0:0].copy(), df2.iloc[0:0].copy()



    '''def _empty_like(self, df):
        """Создает словарь с пустыми значениями для всех колонок DataFrame"""
        return {col: np.nan for col in df.columns}'''

    def _align_pair3(self, df1: pd.DataFrame, df2: pd.DataFrame):
        """
        Линейное выравнивание двух подтаблиц по self.compare_column.
        Построчно, учитывает дубликаты, вставляет заглушки.
        Возвращает два DataFrame одинаковой длины.
        """
        r1, r2 = [], []
        used_j = set()  # индексы уже сопоставленных элементов df2

        for i1, row1 in df1.iterrows():
            val1 = str(row1[self.compare_column]).strip()
            found = False
            for i2, row2 in df2.iterrows():
                if i2 in used_j:
                    continue
                val2 = str(row2[self.compare_column]).strip()
                if val1 == val2:
                    r1.append(row1.to_dict())
                    r2.append(row2.to_dict())
                    used_j.add(i2)
                    found = True
                    break
            if not found:
                r1.append(row1.to_dict())
                r2.append(self._empty_like(df2))

        # Добавляем оставшиеся элементы df2, которые не были сопоставлены
        for i2, row2 in df2.iterrows():
            if i2 not in used_j:
                r1.append(self._empty_like(df1))
                r2.append(row2.to_dict())

        a1 = pd.DataFrame(r1).reindex(columns=df1.columns, fill_value=np.nan)
        a2 = pd.DataFrame(r2).reindex(columns=df2.columns, fill_value=np.nan)

        return a1.reset_index(drop=True), a2.reset_index(drop=True)


    def _align_pair2(self, df1: pd.DataFrame, df2: pd.DataFrame):
        """
        Выравнивает две подтаблицы по self.compare_column.
        Поддерживает дубликаты, вставки, удаления и замены.
        Возвращает два выровненных DataFrame одинаковой длины.
        """
        if df1.empty and df2.empty:
            return df1.copy(), df2.copy()
        if df1.empty:
            r1 = [self._empty_like(df1) for _ in range(len(df2))]
            return pd.DataFrame(r1).reset_index(drop=True), df2.reset_index(drop=True)
        if df2.empty:
            r2 = [self._empty_like(df2) for _ in range(len(df1))]
            return df1.reset_index(drop=True), pd.DataFrame(r2).reset_index(drop=True)

        # Создаем уникальные маркеры для SequenceMatcher (значение + локальный индекс)
        s1 = list(zip(df1[self.compare_column].fillna("").astype(str).str.strip().str.casefold(), range(len(df1))))
        s2 = list(zip(df2[self.compare_column].fillna("").astype(str).str.strip().str.casefold(), range(len(df2))))

        m = SequenceMatcher(None, s1, s2)
        r1, r2 = [], []

        for tag, i1, i2, j1, j2 in m.get_opcodes():
            if tag == "equal":
                r1.extend(df1.iloc[i1:i2].to_dict("records"))
                r2.extend(df2.iloc[j1:j2].to_dict("records"))
            elif tag == "replace":
                for i in range(i1, i2):
                    r1.append(df1.iloc[i].to_dict())
                    r2.append(self._empty_like(df2))
                for j in range(j1, j2):
                    r1.append(self._empty_like(df1))
                    r2.append(df2.iloc[j].to_dict())
            elif tag == "delete":
                for i in range(i1, i2):
                    r1.append(df1.iloc[i].to_dict())
                    r2.append(self._empty_like(df2))
            elif tag == "insert":
                for j in range(j1, j2):
                    r1.append(self._empty_like(df1))
                    r2.append(df2.iloc[j].to_dict())

        a1 = pd.DataFrame(r1).reset_index(drop=True)
        a2 = pd.DataFrame(r2).reset_index(drop=True)

        # Восстановить исходные колонки и порядок
        a1 = a1.reindex(columns=df1.columns, fill_value=np.nan)
        a2 = a2.reindex(columns=df2.columns, fill_value=np.nan)
        return a1, a2

    from difflib import SequenceMatcher
    import pandas as pd
    import numpy as np

    def _align2(self):
        """
        Линейное выравнивание df1 и df2 с учётом подразделов и compare_column.
        Использует SequenceMatcher на кортежах (subsection, compare_column).
        Вставляет заглушки (_empty_like) при отсутствии соответствия.


        """

        df1 = self.df1[self.compare_column].fillna("").astype(str).tolist()
        df2 = self.df2[self.compare_column].fillna("").astype(str).tolist()


        r1, r2 = [], []

        # метка для пустых подразделов
        sentinel_sub = "∅"

        # создаём последовательности кортежей: (subsection, compare_column)
        seq1 = list(zip(
            df1[self.subsection_column].fillna(sentinel_sub),
            df1[self.compare_column].fillna("").astype(str)
        ))
        seq2 = list(zip(
            df2[self.subsection_column].fillna(sentinel_sub),
            df2[self.compare_column].fillna("").astype(str)
        ))

        sm = SequenceMatcher(None, seq1, seq2)

        for tag, i1, i2, j1, j2 in sm.get_opcodes():
            if tag == "equal":
                r1.extend(df1.iloc[i1:i2].to_dict("records"))
                r2.extend(df2.iloc[j1:j2].to_dict("records"))

            elif tag in ("replace", "delete", "insert"):
                # строки из df1
                for i in range(i1, i2):
                    r1.append(df1.iloc[i].to_dict())
                    if j1 < j2:
                        dummy = df2.iloc[j1].copy().to_dict()
                        if self.value_column in dummy:
                            dummy[self.value_column] = 0
                        r2.append(dummy)
                        j1 += 1
                    else:
                        if not df2.empty:
                            dummy = df2.iloc[0].copy().to_dict()
                            if self.value_column in dummy:
                                dummy[self.value_column] = 0
                            r2.append(dummy)
                        else:
                            r2.append({col: None for col in df2.columns})

                # строки из df2, которые остались
                for j in range(j1, j2):
                    if not df1.empty:
                        dummy = df1.iloc[0].copy().to_dict()
                        if self.value_column in dummy:
                            dummy[self.value_column] = 0
                        r1.append(dummy)
                    else:
                        r1.append({col: None for col in df1.columns})
                    r2.append(df2.iloc[j].to_dict())

        # создаём DataFrame и восстанавливаем порядок колонок
        a1 = pd.DataFrame(r1, dtype=object).reindex(columns=df1.columns)
        a2 = pd.DataFrame(r2, dtype=object).reindex(columns=df2.columns)
        return a1.reset_index(drop=True), a2.reset_index(drop=True)

    def _align(self):
        s1 = self._build_alignment_key_series(self.df1).tolist()
        s2 = self._build_alignment_key_series(self.df2).tolist()

        m = SequenceMatcher(None, s1, s2)
        r1, r2 = [], []
        for tag, i1, i2, j1, j2 in m.get_opcodes():
            if tag == "equal":
                r1.extend(self.df1.iloc[i1:i2].to_dict("records"))
                r2.extend(self.df2.iloc[j1:j2].to_dict("records"))
            elif tag == "replace":
                for i in range(i1, i2):
                    r1.append(self.df1.iloc[i].to_dict())
                    r2.append(self._empty_like(self.df2))
                for j in range(j1, j2):
                    r1.append(self._empty_like(self.df1))
                    r2.append(self.df2.iloc[j].to_dict())
            elif tag == "delete":
                for i in range(i1, i2):
                    r1.append(self.df1.iloc[i].to_dict())
                    r2.append(self._empty_like(self.df2))
            elif tag == "insert":
                for j in range(j1, j2):
                    r1.append(self._empty_like(self.df1))
                    r2.append(self.df2.iloc[j].to_dict())
        return pd.DataFrame(r1).reset_index(drop=True), pd.DataFrame(r2).reset_index(drop=True)


    def _align_art(self):
        df1 = self.df1.copy()
        df2 = self.df2.copy()
        s1 = self._build_alignment_key_series(df1).tolist()
        s2 = self._build_alignment_key_series(df2).tolist()

        m = SequenceMatcher(None, s1, s2)

        r1, r2 = [], []
        for tag, i1, i2, j1, j2 in m.get_opcodes():
            if tag == "equal":
                r1.extend(df1.iloc[i1:i2].to_dict("records"))
                r2.extend(df2.iloc[j1:j2].to_dict("records"))
            elif tag == "replace":
                for i in range(i1, i2):
                    r1.append(df1.iloc[i].to_dict())
                    r2.append(self._empty_like(df2))
                for j in range(j1, j2):
                    r1.append(self._empty_like(df1))
                    r2.append(df2.iloc[j].to_dict())
            elif tag == "delete":
                for i in range(i1, i2):
                    r1.append(df1.iloc[i].to_dict())
                    r2.append(self._empty_like(df2))
            elif tag == "insert":
                for j in range(j1, j2):
                    r1.append(self._empty_like(df1))
                    r2.append(df2.iloc[j].to_dict())
        df_r1 = pd.DataFrame(r1)
        df_r2 = pd.DataFrame(r2)

        all_cols = sorted(set(df_r1.columns) | set(df_r2.columns))
        df_r1 = df_r1.reindex(columns=all_cols)
        df_r2 = df_r2.reindex(columns=all_cols)

        return df_r1.reset_index(drop=True), df_r2.reset_index(drop=True)

        '''return pd.DataFrame(r1).reset_index(drop=True), pd.DataFrame(r2).reset_index(drop=True)'''

    from difflib import SequenceMatcher
    import pandas as pd

    def _align_with_subsection(self):
        col_sub = self.subsection_column
        col_cmp = self.compare_column
        tmp_col = "__key__"

        # Делаем копии, чтобы не портить исходные df
        df1 = self.df1.copy()
        df2 = self.df2.copy()

        # Создаём временный ключ
        df1[tmp_col] = df1[col_sub].fillna("").astype(str) + "||" + df1[col_cmp].fillna("").astype(str)
        df2[tmp_col] = df2[col_sub].fillna("").astype(str) + "||" + df2[col_cmp].fillna("").astype(str)

        # Меняем compare_column на временный
        old_cmp = self.compare_column
        self.compare_column = tmp_col

        # Запускаем старый _align
        a1, a2 = self._align()

        # Возвращаем compare_column обратно
        self.compare_column = old_cmp

        # Убираем временный ключ
        if tmp_col in a1.columns:
            a1 = a1.drop(columns=[tmp_col])
        if tmp_col in a2.columns:
            a2 = a2.drop(columns=[tmp_col])

        return a1, a2

    def _empty_like2(df: pd.DataFrame) -> dict:
        """Создать пустую строку-словарь для DataFrame."""
        return {col: None for col in df.columns}

    def _empty_like(self, df: pd.DataFrame) -> dict:
        """Создает пустую строку-словарь с теми же колонками, что у df"""
        return {col: "" for col in df.columns}

    def _align7(df1: pd.DataFrame, df2: pd.DataFrame, compare_column: str):
        """
        Линейное выравнивание двух DataFrame по столбцу compare_column.
        Использует SequenceMatcher, поддерживает вставки/удаления.
        """
        s1 = df1[compare_column].fillna("").astype(str).tolist()
        s2 = df2[compare_column].fillna("").astype(str).tolist()

        m = SequenceMatcher(None, s1, s2)
        r1, r2 = [], []
        for tag, i1, i2, j1, j2 in m.get_opcodes():
            if tag == "equal":
                r1.extend(df1.iloc[i1:i2].to_dict("records"))
                r2.extend(df2.iloc[j1:j2].to_dict("records"))
            elif tag == "replace":
                for i in range(i1, i2):
                    r1.append(df1.iloc[i].to_dict())
                    r2.append(_empty_like(df2))
                for j in range(j1, j2):
                    r1.append(_empty_like(df1))
                    r2.append(df2.iloc[j].to_dict())
            elif tag == "delete":
                for i in range(i1, i2):
                    r1.append(df1.iloc[i].to_dict())
                    r2.append(_empty_like(df2))
            elif tag == "insert":
                for j in range(j1, j2):
                    r1.append(_empty_like(df1))
                    r2.append(df2.iloc[j].to_dict())

        return (
            pd.DataFrame(r1).reset_index(drop=True),
            pd.DataFrame(r2).reset_index(drop=True))

    _SECTION_NAME_RE = re.compile(r'^\s*раздел\s+\d+\.?\s*', re.IGNORECASE)

    def _strip_section(self, text: str | None) -> str:
        """Убирает префикс 'Раздел N. ' и лишние пробелы."""
        return self._SECTION_NAME_RE.sub('', (text or '')).strip()

    def _norm(self, text: str | None) -> str:
        return (text or '').strip().lower()

    def _section_match_series(self, df: pd.DataFrame) -> pd.Series:
        if "Название раздела" in df.columns:
            return df["Название раздела"].fillna("").astype(str).str.strip()
        if "Раздел" in df.columns:
            return df["Раздел"].map(self._strip_section)
        return pd.Series([""] * len(df), index=df.index, dtype="object")

    def _build_alignment_key_series(self, df: pd.DataFrame) -> pd.Series:
        section = self._section_match_series(df)
        if self.subsection_column and self.subsection_column in df.columns:
            subsection = df[self.subsection_column].fillna("").astype(str).str.strip()
        else:
            subsection = pd.Series([""] * len(df), index=df.index, dtype="object")
        compare = df[self.compare_column].fillna("").astype(str).str.strip()
        return section + "||" + subsection + "||" + compare
    # subsection rows
    def _insert_subsection(self, rows: List[dict], subs1: pd.Series, subs2: pd.Series, col_order: List[str]):
        current = None
        out: List[dict] = []
        for i, r in enumerate(rows):
            sub = subs1[i] or subs2[i]
            if sub and sub != current:
                divider = {c: (np.nan if c not in (self.compare_column, *self.extra_column) else "") for c in col_order}
                divider[self.compare_column] = f"-- {sub} --"
                out.append(divider)
                current = sub
            out.append(r)
        return out

    # number formatting
    @staticmethod
    def _fmt_num(x):
        if x in ("", None) or (isinstance(x, float) and np.isnan(x)):
            return ""
        try:
            return f"{float(str(x).replace(' ', '').replace(',', '.')):,.2f}".replace(",", " ")
        except ValueError:
            return ""

    def _fmt_diff(self, x):
        if x in ("", None) or (isinstance(x, float) and np.isnan(x)):
            return ""
        try:
            v = float(str(x).replace(" ", "").replace(",", "."))
            f = f"{v:,.2f}".replace(",", " ")
            if v > 0:
                return f"<span class='diff-positive'>{f}</span>"
            elif v < 0:
                return f"<span class='diff-negative'>{f}</span>"
            return f
        except ValueError:
            return x

    # coalesce text
    @staticmethod
    def _coalesce(a: pd.Series, b: pd.Series):
        return a.fillna("").astype(str).replace("", np.nan).combine_first(b.astype(str)).fillna("")

    # ------------------------------------------------------------------
    # Customer report
    # ------------------------------------------------------------------
    # ────────────────────────────────────────────────────────────────
    #  helpers (приватные)
    # ────────────────────────────────────────────────────────────────
    def _norm(self, text: str | None) -> str:
        """Нормализуем строку для сравнения: обрезаем пробелы, lower-case."""
        return (text or "").strip().lower()

    # ------------------------------------------------------------------
    #  Customer-report  (Раздел  ▸ Подраздел ▸ позиции)
    # ------------------------------------------------------------------
    def generate_subsection_summary(self) -> pd.DataFrame:
        """
        Сводка по подразделам: Раздел | Подраздел | Стоимость (file1) |
        Стоимость (file2) | Разница. Всегда использует колонку "Стоимость".
        """
        # 1) Выравниваем две таблицы
        #d1, d2 = self._align()
        d1, d2 = self._align_art()
        #print('пробуем')
        key = self.subsection_column

        # 2) Жестко берем колонку "Стоимость" для агрегирования
        c = "Стоимость"
        #print('#1')

        # 3) Подготовка названий разделов без префиксов
        if "Раздел" in d1.columns:
            #print('#2')
            d1 = d1.copy();
            d2 = d2.copy()
            #print('#3')
            if "Название раздела" in d1.columns and "Название раздела" in d2.columns:
                d1["SectionName"] = d1["Название раздела"].fillna("").astype(str).str.strip()
                d2["SectionName"] = d2["Название раздела"].fillna("").astype(str).str.strip()
            else:
                d1["SectionName"] = d1["Раздел"].map(self._strip_section)
                d2["SectionName"] = d2["Раздел"].map(self._strip_section)
            #print('#4')
            grp_cols = ["SectionName", key]
            #print('#5')
        else:
            grp_cols = [key]
            #print('#6')

        # 4) Группировка и суммирование по подразделам
        grp1 = (
            d1.groupby(grp_cols, dropna=False)[c]
            .sum().reset_index()
            .rename(columns={c: f"{c} ({self.file1_name})"})
        )
        grp2 = (
            d2.groupby(grp_cols, dropna=False)[c]
            .sum().reset_index()
            .rename(columns={c: f"{c} ({self.file2_name})"})
        )
        #('#7')

        # 5) Объединяем и вычисляем разницу
        summary = grp1.merge(grp2, on=grp_cols, how="outer").fillna(0)
        #print('#8')
        col1 = f"{c} ({self.file1_name})"
        col2 = f"{c} ({self.file2_name})"
        #print('#9')
        summary[col1] = pd.to_numeric(summary[col1], errors="coerce").fillna(0)
        summary[col2] = pd.to_numeric(summary[col2], errors="coerce").fillna(0)
        summary[f"Разница ({c})"] = summary[col1] - summary[col2]

        #print('#10')

        # 6) Восстанавливаем исходный порядок появления (d1 потом d2)
        order = []
        for sec, sub in zip(d1.get("SectionName", []), d1.get(key, [])):
            pair = (sec, sub)
            if pair not in order:
                order.append(pair)
        for sec, sub in zip(d2.get("SectionName", []), d2.get(key, [])):
            pair = (sec, sub)
            if pair not in order:
                order.append(pair)
        order_index = {pair: idx for idx, pair in enumerate(order)}
        summary["_order"] = summary.apply(
            lambda r: order_index.get((r.get("SectionName", r.get("Раздел")), r[key]), len(order)),
            axis=1
        )
        #print('#11')
        summary = (
            summary.sort_values("_order")
            .drop(columns=["_order"])
            .reset_index(drop=True)
        )
        #print('#12')

        # 7) Переименовываем SectionName обратно в Раздел, если нужно
        if "SectionName" in summary.columns:
            summary = summary.rename(columns={"SectionName": "Раздел"})

        #print('#13')

        return summary

    def generate_customer_report(self) -> pd.DataFrame:
        d1, d2 = self._align_art()
        d1[self.value_column] = d1[self.value_column].apply(pd.to_numeric, errors="coerce").fillna(0)
        d2[self.value_column] = d2[self.value_column].apply(pd.to_numeric, errors="coerce").fillna(0)

        #print(d1.shape, d2.shape)
        #print('probuem')
        rows: list[dict] = []
        cur_sec = cur_sub = None
        counter = 1

        has_sec = "Раздел" in d1.columns
        has_sub = self.subsection_column and self.subsection_column in d1.columns
        #print('probuem2')
        for i in range(len(d1)):
            # 1) raw
            #print('#1')
            sec_source = "Название раздела" if "Название раздела" in d1.columns and "Название раздела" in d2.columns else "Раздел"
            raw_sec = (d1[sec_source][i] if has_sec else "") or (d2[sec_source][i] if has_sec else "")
            sec = raw_sec.strip() if sec_source == "Название раздела" else self._strip_section(raw_sec)
            #print('#2')

            raw_sub = (d1[self.subsection_column][i] if has_sub else "") or \
                      (d2[self.subsection_column][i] if has_sub else "")
            sub     = raw_sub.strip()
            #print('#3')

            # 2) новый раздел?
            if sec and self._norm(sec) != self._norm(cur_sec):
                print('#4')
                rows.append(self._divider_row(f"-- {sec} --"))
                cur_sec, cur_sub = sec, None

            # 3) новый подраздел? и не совпадает с разделом
            if sub and self._norm(sub) != self._norm(cur_sub) and self._norm(sub) != self._norm(cur_sec):
                #print('#5')
                rows.append(self._divider_row(f"---- {sub} ----"))
                cur_sub = sub

            # 4) позиция
            row = {"№": counter}
            counter += 1
            #print('#6')

            # compare_column
            row[self.compare_column] = self._coalesce(
                d1[self.compare_column][i:i+1],
                d2[self.compare_column][i:i+1]
            ).iloc[0]

            # extra
            for col in self.extra_column:
                #print('#7')
                row[col] = self._coalesce(
                    d1.get(col, pd.Series([""]*len(d1)))[i:i+1],
                    d2.get(col, pd.Series([""]*len(d2)))[i:i+1]
                ).iloc[0]



            # numeric + diff
            for v in self.value_column:
                #print('#8')
                #print(v)
                a, b = d1[v][i], d2[v][i]
                #print(a, b)
                row[f"{v} ({self.file1_name})"] = a
                row[f"{v} ({self.file2_name})"] = b
                #print('#8.1')
                row[f"Разница ({v})"] = a - b
                #print('#9')

            #print('#10')

            rows.append(row)
        #print('probuem3')
        return pd.DataFrame(rows)

    def generate_top_difference_report(self, coverage: float = 0.8) -> pd.DataFrame:
        d1, d2 = self._align_art()

        qty1 = pd.to_numeric(d1.get("Количество", pd.Series([None] * len(d1))), errors="coerce")
        qty2 = pd.to_numeric(d2.get("Количество", pd.Series([None] * len(d2))), errors="coerce")
        cost1 = pd.to_numeric(d1.get("Стоимость", pd.Series([None] * len(d1))), errors="coerce").fillna(0)
        cost2 = pd.to_numeric(d2.get("Стоимость", pd.Series([None] * len(d2))), errors="coerce").fillna(0)
        pos1 = d1.get("Номер позиции", pd.Series([""] * len(d1))).fillna("").astype(str).str.strip()
        pos2 = d2.get("Номер позиции", pd.Series([""] * len(d2))).fillna("").astype(str).str.strip()
        name1 = d1.get(self.compare_column, pd.Series([""] * len(d1))).fillna("").astype(str).str.strip()
        name2 = d2.get(self.compare_column, pd.Series([""] * len(d2))).fillna("").astype(str).str.strip()
        section1 = d1.get("Название раздела", d1.get("Раздел", pd.Series([""] * len(d1)))).fillna("").astype(str).str.strip()
        section2 = d2.get("Название раздела", d2.get("Раздел", pd.Series([""] * len(d2)))).fillna("").astype(str).str.strip()
        subsection1 = d1.get("Подраздел", pd.Series([""] * len(d1))).fillna("").astype(str).str.strip()
        subsection2 = d2.get("Подраздел", pd.Series([""] * len(d2))).fillna("").astype(str).str.strip()

        info = pd.DataFrame({
            "Раздел": section1.where(section1 != "", section2),
            "Подраздел": subsection1.where(subsection1 != "", subsection2),
            "№": pos1.where(pos1 != "", pos2),
            self.compare_column: name1.where(name1 != "", name2),
            "Кол-во (Проект)": qty1,
            "Кол-во (Факт)": qty2,
            "Ст-ть (Проект)": cost1,
            "Ст-ть (Факт)": cost2,
        })
        info["Разница (Ст-ть)"] = info["Ст-ть (Проект)"] - info["Ст-ть (Факт)"]
        info = info[info["Разница (Ст-ть)"] != 0].copy()

        def select_by_coverage(frame: pd.DataFrame, positive: bool) -> pd.DataFrame:
            subset = frame[frame["Разница (Ст-ть)"] > 0].copy() if positive else frame[frame["Разница (Ст-ть)"] < 0].copy()
            if subset.empty:
                return subset
            subset["_weight"] = subset["Разница (Ст-ть)"] if positive else subset["Разница (Ст-ть)"].abs()
            subset = subset.sort_values("_weight", ascending=False)
            total_weight = subset["_weight"].sum()
            if total_weight <= 0:
                return subset.iloc[:0]
            subset["_cum_share"] = subset["_weight"].cumsum() / total_weight
            selected = subset[subset["_cum_share"] <= coverage].copy()
            min_rows = min(3, len(subset))
            if len(selected) < min_rows:
                selected = subset.head(min_rows).copy()
            elif len(selected) < len(subset):
                next_row = subset.iloc[[len(selected)]].copy()
                selected = pd.concat([selected, next_row], ignore_index=True)
            return selected.drop(columns=["_weight", "_cum_share"], errors="ignore")

        positive_rows = select_by_coverage(info, positive=True)
        negative_rows = select_by_coverage(info, positive=False)
        result = pd.concat([positive_rows, negative_rows], ignore_index=True)
        result["_abs_diff"] = result["Разница (Ст-ть)"].abs()
        result = result.sort_values(["_abs_diff", "Разница (Ст-ть)"], ascending=[False, False])
        return result.drop(columns=["_abs_diff"], errors="ignore").reset_index(drop=True)

    def generate_unit_difference_report(self) -> pd.DataFrame:
        d1, d2 = self._align_art()

        qty1 = pd.to_numeric(d1.get("Количество", pd.Series([None] * len(d1))), errors="coerce")
        qty2 = pd.to_numeric(d2.get("Количество", pd.Series([None] * len(d2))), errors="coerce")
        cost1 = pd.to_numeric(d1.get("Стоимость", pd.Series([None] * len(d1))), errors="coerce").fillna(0)
        cost2 = pd.to_numeric(d2.get("Стоимость", pd.Series([None] * len(d2))), errors="coerce").fillna(0)
        pos1 = d1.get("Номер позиции", pd.Series([""] * len(d1))).fillna("").astype(str).str.strip()
        pos2 = d2.get("Номер позиции", pd.Series([""] * len(d2))).fillna("").astype(str).str.strip()
        name1 = d1.get(self.compare_column, pd.Series([""] * len(d1))).fillna("").astype(str).str.strip()
        name2 = d2.get(self.compare_column, pd.Series([""] * len(d2))).fillna("").astype(str).str.strip()
        section1 = d1.get("Название раздела", d1.get("Раздел", pd.Series([""] * len(d1)))).fillna("").astype(str).str.strip()
        section2 = d2.get("Название раздела", d2.get("Раздел", pd.Series([""] * len(d2)))).fillna("").astype(str).str.strip()
        subsection1 = d1.get("Подраздел", pd.Series([""] * len(d1))).fillna("").astype(str).str.strip()
        subsection2 = d2.get("Подраздел", pd.Series([""] * len(d2))).fillna("").astype(str).str.strip()
        unit1 = d1.get("Единица измерения", pd.Series([""] * len(d1))).fillna("").astype(str).str.strip()
        unit2 = d2.get("Единица измерения", pd.Series([""] * len(d2))).fillna("").astype(str).str.strip()

        report = pd.DataFrame({
            "Раздел": section1.where(section1 != "", section2),
            "Подраздел": subsection1.where(subsection1 != "", subsection2),
            "№": pos1.where(pos1 != "", pos2),
            self.compare_column: name1.where(name1 != "", name2),
            "Ед. изм. (Проект)": unit1,
            "Ед. изм. (Факт)": unit2,
            "Кол-во (Проект)": qty1,
            "Кол-во (Факт)": qty2,
            "Ст-ть (Проект)": cost1,
            "Ст-ть (Факт)": cost2,
        })
        report["Разница (Ст-ть)"] = report["Ст-ть (Проект)"] - report["Ст-ть (Факт)"]
        mismatch_mask = (
            (report["Ед. изм. (Проект)"] != report["Ед. изм. (Факт)"])
            & (report["Ед. изм. (Проект)"] != "")
            & (report["Ед. изм. (Факт)"] != "")
        )
        report = report[mismatch_mask].copy()
        report["_abs_diff"] = report["Разница (Ст-ть)"].abs()
        report = report.sort_values(["_abs_diff", "№"], ascending=[False, True])
        return report.drop(columns=["_abs_diff"]).reset_index(drop=True)

    # ------------------------------------------------------------------
    #  вспомогательный «разделитель»
    # ------------------------------------------------------------------
    _SECTION_NAME_RE = re.compile(r'^\s*раздел\s+\d+\.?\s*', re.IGNORECASE)

    def _strip_section(self, text: str | None) -> str:
        """Убирает префикс 'Раздел N. ' и лишние пробелы."""
        return self._SECTION_NAME_RE.sub('', (text or '')).strip()

    def _norm(self, text: str | None) -> str:
        return (text or '').strip().lower()

    def _divider_row(self, text: str) -> dict:
        row = {"№": "", self.compare_column: text}
        for col in (*self.extra_column,
                    *(f"{v} ({self.file1_name})" for v in self.value_column),
                    *(f"{v} ({self.file2_name})" for v in self.value_column),
                    *(f"Разница ({v})" for v in self.value_column)):
            row.setdefault(col, "")
        return row

    @staticmethod
    def _is_divider_text(value) -> bool:
        return isinstance(value, str) and value.strip().startswith("--")

    def export_customer_html(self, path: str = "customer_report.html") -> str:
        """
        Экспортит в один HTML-файл две таблицы подряд:
          1) Детальный отчёт
          2) Сводка по подразделам
        """
        # 1. Сгенерировать оба DataFrame
        df_detail = self.generate_customer_report()
        df_summary = self.generate_subsection_summary()

        # 2. Универсальное CSS для обеих таблиц
        css = """
        <style>
          table { width:100%; border-collapse:collapse; }
          th, td { padding:6px 10px; border:1px solid #ccc; text-align:left; }
          th { background:#f5f5f5; }
          .diff-positive { color:green; font-weight:bold; }
          .diff-negative { color:red; font-weight:bold; }
          .divider { background:#E9ECEF; font-weight:bold; }
        </style>
        """

        # 3. Подсветка строк-разделителей в детальном отчёте
        def highlight_dividers(html: str) -> str:
            return re.sub(r"<tr><td>(--+ .+? --+)</td>", r'<tr class="divider"><td>\1</td>', html)

        # 4. Сборка итогового HTML
        with open(path, "w", encoding="utf-8") as f:
            f.write("<html><head><meta charset='utf-8'>")
            f.write(css)
            f.write("</head><body>")

            # Детальный отчёт
            f.write("<h2>Детальный отчёт для заказчика</h2>")
            detail_html = df_detail.to_html(index=False, escape=False)
            f.write(highlight_dividers(detail_html))

            # Сводка по подразделам
            f.write("<h2>Сводка по подразделам</h2>")
            f.write(df_summary.to_html(index=False, escape=False))

            f.write("</body></html>")

        return path

    def export_customer_excel(
            self,
            path: str = "customer_report.xlsx",
    ) -> str:
        """
        Экспортит в один Excel-файл два листа:
          – 'Customer' : детальный отчёт (с раскраской, если есть Категория)
          – 'Summary'  : сводка по подразделам
        """
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import PatternFill, Font, Border, Side

        # 1) Получаем DataFrame
        df_detail = self.generate_customer_report()
        df_summary = self.generate_subsection_summary()
        df_info = self.generate_top_difference_report()
        df_unit_diff = self.generate_unit_difference_report()
        short_names = {
            "Количество": "Кол-во",
            "Стоимость": "Ст-ть",
        }

        def shorten_metric(name: str) -> str:
            return short_names.get(name, name)

        def with_line_suffix(label: str, suffix: str) -> str:
            return f"{label}\n({suffix})"

        detail_renames = {}
        summary_renames = {}
        for value_name in self.value_column:
            short_value = shorten_metric(value_name)
            detail_renames[f"{value_name} ({self.file1_name})"] = with_line_suffix(short_value, "Проект")
            detail_renames[f"{value_name} ({self.file2_name})"] = with_line_suffix(short_value, "Факт")
            detail_renames[f"Разница ({value_name})"] = f"Разница\n({short_value})"
        summary_renames[f"Стоимость ({self.file1_name})"] = with_line_suffix("Ст-ть", "Проект")
        summary_renames[f"Стоимость ({self.file2_name})"] = with_line_suffix("Ст-ть", "Факт")
        summary_renames["Разница (Стоимость)"] = "Разница\n(Ст-ть)"

        export_detail = df_detail.rename(columns=detail_renames)
        export_summary = df_summary.rename(columns=summary_renames)
        export_detail = export_detail.rename(columns={"Единица измерения": "Ед.изм."})
        df_info = df_info.rename(columns={
            "Кол-во (Проект)": with_line_suffix("Кол-во", "Проект"),
            "Кол-во (Факт)": with_line_suffix("Кол-во", "Факт"),
            "Ст-ть (Проект)": with_line_suffix("Ст-ть", "Проект"),
            "Ст-ть (Факт)": with_line_suffix("Ст-ть", "Факт"),
            "Разница (Ст-ть)": "Разница\n(Ст-ть)",
            "Ед. изм. (Проект)": with_line_suffix("Ед.изм.", "Проект"),
            "Ед. изм. (Факт)": with_line_suffix("Ед.изм.", "Факт"),
        })
        df_unit_diff = df_unit_diff.rename(columns={
            "Кол-во (Проект)": with_line_suffix("Кол-во", "Проект"),
            "Кол-во (Факт)": with_line_suffix("Кол-во", "Факт"),
            "Ст-ть (Проект)": with_line_suffix("Ст-ть", "Проект"),
            "Ст-ть (Факт)": with_line_suffix("Ст-ть", "Факт"),
            "Разница (Ст-ть)": "Разница\n(Ст-ть)",
            "Ед. изм. (Проект)": with_line_suffix("Ед.изм.", "Проект"),
            "Ед. изм. (Факт)": with_line_suffix("Ед.изм.", "Факт"),
        })
        preferred_order = [
            "№",
            "Код расценки",
            "Наименование",
            "Ед.изм.",
            "Раздел",
            "Подраздел",
            "Кол-во\n(Проект)",
            "Кол-во\n(Факт)",
            "Разница\n(Кол-во)",
            "Ст-ть\n(Проект)",
            "Ст-ть\n(Факт)",
            "Разница\n(Ст-ть)",
            "Категория",
        ]
        detail_columns = [col for col in preferred_order if col in export_detail.columns]
        detail_columns.extend(col for col in export_detail.columns if col not in detail_columns)
        export_detail = export_detail[detail_columns]

        # 2) Подготовка стилей
        fill_work = PatternFill("solid", fgColor="D6EFD6")
        fill_mat = PatternFill("solid", fgColor="D6EAF8")
        fill_red = PatternFill("solid", fgColor="F8D6D6")
        fill_div = PatternFill("solid", fgColor="E9ECEF")
        bold = Font(bold=True)
        thin = Side(style="thin", color="000000")
        border_b = Border(bottom=thin)
        money_format = '#,##0.00'

        # 3) Запись в Excel
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            # — Детальный лист —
            export_detail.to_excel(writer, index=False, sheet_name="Customer")
            ws = writer.sheets["Customer"]

            headers = list(export_detail.columns)
            # ищем нужные индексы
            idx_cat = headers.index("Категория") + 1 if "Категория" in headers else None
            idx_qty = headers.index("Кол-во\n(Проект)") + 1 if "Кол-во\n(Проект)" in headers else None
            idx_compare = headers.index(self.compare_column) + 1 if self.compare_column in headers else None

            # проходим по строкам
            for r in range(2, len(export_detail) + 2):
                divider_value = ws.cell(row=r, column=idx_compare).value if idx_compare else None

                # строка-разделитель?
                if self._is_divider_text(divider_value):
                    for c in range(1, len(headers) + 1):
                        cell = ws.cell(row=r, column=c)
                        cell.fill = fill_div
                        cell.font = bold
                    continue

                # раскраска по Категории (если есть)
                cat = None
                if idx_cat:
                    cat = ws.cell(row=r, column=idx_cat).value

                qty = None
                if idx_qty:
                    try:
                        qty = ws.cell(row=r, column=idx_qty).value
                    except:
                        qty = None

                if cat == "Работа":
                    fill = fill_work
                elif cat == "Материалы":
                    fill = fill_red if (isinstance(qty, (int, float)) and qty < 0) else fill_mat
                else:
                    fill = None

                if fill:
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=r, column=c).fill = fill

                # граница над строкой, если текущая категория "Работа" или "Механизмы"
                if cat in ("Работа", "Механизмы"):
                    for c in range(1, len(headers) + 1):
                        ws.cell(row=r - 1, column=c).border = border_b

            apply_readable_sheet_layout(ws, export_detail, max_row_height=None)
            apply_section_row_grouping(ws, export_detail, self.compare_column)
            autofit_columns_by_name(
                ws,
                export_detail,
                [f"Кол-во\n(Проект)", f"Кол-во\n(Факт)", f"Ст-ть\n(Проект)", f"Ст-ть\n(Факт)"],
                min_width=12,
                max_width=24,
                wrap_lines=2,
            )
            if "Код расценки" in headers:
                i = headers.index("Код расценки") + 1
                ws.column_dimensions[get_column_letter(i)].width = 20
            if "Категория" in headers:
                i = headers.index("Категория") + 1
                ws.column_dimensions[get_column_letter(i)].width = 10
            if "Раздел" in headers:
                section_col = headers.index("Раздел") + 1
                for row_index in range(2, len(export_detail) + 2):
                    cell = ws.cell(row=row_index, column=section_col)
                    if str(cell.value or "").strip():
                        cell.font = bold
            for col_index, header in enumerate(headers, start=1):
                if "Ст-ть" not in str(header) and "Разница (Стоимость)" not in str(header):
                    continue
                for row_index in range(2, len(export_detail) + 2):
                    cell = ws.cell(row=row_index, column=col_index)
                    if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                        cell.number_format = money_format

            # — Сводный лист —
            export_summary.to_excel(writer, index=False, sheet_name="Summary")
            summary_ws = writer.sheets["Summary"]
            apply_readable_sheet_layout(summary_ws, export_summary)
            for col_index, header in enumerate(export_summary.columns, start=1):
                if "Ст-ть" not in str(header) and "Разница" not in str(header):
                    continue
                for row_index in range(2, len(export_summary) + 2):
                    cell = summary_ws.cell(row=row_index, column=col_index)
                    if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                        cell.number_format = money_format

            df_info.to_excel(writer, index=False, sheet_name="Инфо")
            info_ws = writer.sheets["Инфо"]
            apply_readable_sheet_layout(info_ws, df_info)
            apply_named_column_widths(
                info_ws,
                list(df_info.columns),
                {
                    "Раздел": 26,
                    "Подраздел": 28,
                    "№": 14,
                    "Наименование": 60,
                    "Разница\n(Ст-ть)": 16,
                },
            )
            autofit_columns_by_name(
                info_ws,
                df_info,
                ["Кол-во\n(Проект)", "Кол-во\n(Факт)", "Ст-ть\n(Проект)", "Ст-ть\n(Факт)"],
                min_width=12,
                max_width=24,
                wrap_lines=2,
            )
            cap_row_heights(info_ws, 36)
            for col_index, header in enumerate(df_info.columns, start=1):
                if "Ст-ть" not in str(header):
                    continue
                for row_index in range(2, len(df_info) + 2):
                    cell = info_ws.cell(row=row_index, column=col_index)
                    if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                        cell.number_format = money_format

            df_unit_diff.to_excel(writer, index=False, sheet_name="Отличается единица измерения")
            unit_ws = writer.sheets["Отличается единица измерения"]
            apply_readable_sheet_layout(unit_ws, df_unit_diff)
            apply_named_column_widths(
                unit_ws,
                list(df_unit_diff.columns),
                {
                    "Раздел": 26,
                    "Подраздел": 28,
                    "№": 14,
                    "Наименование": 52,
                    "Ед.изм.\n(Проект)": 16,
                    "Ед.изм.\n(Факт)": 16,
                    "Разница\n(Ст-ть)": 16,
                },
            )
            autofit_columns_by_name(
                unit_ws,
                df_unit_diff,
                ["Кол-во\n(Проект)", "Кол-во\n(Факт)", "Ст-ть\n(Проект)", "Ст-ть\n(Факт)"],
                min_width=12,
                max_width=24,
                wrap_lines=2,
            )
            for col_index, header in enumerate(df_unit_diff.columns, start=1):
                if "Ст-ть" not in str(header):
                    continue
                for row_index in range(2, len(df_unit_diff) + 2):
                    cell = unit_ws.cell(row=row_index, column=col_index)
                    if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                        cell.number_format = money_format

            legend = pd.DataFrame(
                [
                    {
                        "Роль": "Проект",
                        "Файл": self.file1_name,
                        "Общая стоимость": float(pd.to_numeric(self.df1.get("Стоимость"), errors="coerce").fillna(0).sum()),
                    },
                    {
                        "Роль": "Факт",
                        "Файл": self.file2_name,
                        "Общая стоимость": float(pd.to_numeric(self.df2.get("Стоимость"), errors="coerce").fillna(0).sum()),
                    },
                ]
            )
            legend.to_excel(writer, index=False, sheet_name="Файлы")
            files_ws = writer.sheets["Файлы"]
            apply_readable_sheet_layout(files_ws, legend, name_column="Файл", no_wrap_columns={"Файл"})
            autofit_columns_by_name(files_ws, legend, ["Файл"], min_width=30, max_width=120, wrap_lines=1)
            if "Общая стоимость" in legend.columns:
                cost_col = list(legend.columns).index("Общая стоимость") + 1
                for row_index in range(2, len(legend) + 2):
                    cell = files_ws.cell(row=row_index, column=cost_col)
                    if isinstance(cell.value, (int, float)) and not pd.isna(cell.value):
                        cell.number_format = money_format

        return path

    # Missing positions (present in d1, absent in d2)
    # ------------------------------------------------------------------
    def export_positions_absent_in_d2(
            self,
            txt_path: str = "missing_in_d2.txt",
            value_col: Optional[str] = None,
            order: list[str] | None = None,  # порядок не используется, но для единообразия
    ) -> List[str]:
        """Позиции, присутствующие в *df1*, но отсутствующие в *df2* **и** не
        имеющие отрицательной стоимости в *df1*.

        Parameters
        ----------
        txt_path : str
            Куда сохранить TXT‑файл (по строке на позицию).
        value_col : str | None
            Какая колонка стоимости используется для проверки отрицательных
            значений. Если `None`, берётся первый из ``self.value_column``.
        """
        if value_col is None:
            value_col = self.value_column[0]

        d1, d2 = self._align()
        # отсутствует в d2
        absent_mask = d2[self.compare_column].isna() | (
            d2[self.compare_column].astype(str).str.strip() == ""
        )
        # и НЕ отрицательное значение в d1 (NaN/положительное/ноль допускаем)
        non_negative_mask = d1[value_col].isna() | (d1[value_col] >= 0)

        final_mask = absent_mask & non_negative_mask
        missing_list = d1.loc[final_mask, self.compare_column].astype(str).tolist()

        # запись в файл
        with open(txt_path, "w", encoding="utf-8") as f:
            for pos in missing_list:
                f.write(f"{pos}")

        # вывод на экран
        print("*** Позиции из d1, отсутствующие в d2 (стоимость неотрицательная) ***")
        for pos in missing_list:
            print(pos)

        return missing_list

    def get_missing_positions(
            self,
            value_col: Optional[str] = None
    ) -> List[str]:
        """
        Возвращает список позиций из df1, отсутствующих в df2
        (и с неотрицательной стоимостью в df1), без записи в файл.
        """
        if value_col is None:
            value_col = self.value_column[0]

        d1, d2 = self._align()
        absent_mask = d2[self.compare_column].isna() | (
                d2[self.compare_column].astype(str).str.strip() == ""
        )
        nonneg_mask = d1[value_col].isna() | (d1[value_col] >= 0)
        final_mask = absent_mask & nonneg_mask
        return d1.loc[final_mask, self.compare_column].astype(str).tolist()

    def get_added_removed_positions(
            self,
            value_col: Optional[str] = None
    ) -> tuple[list[str], list[str]]:
        """
        Возвращает два списка:
          - исключаемые: позиции из df1, отсутствующие в df2 и со значением > 0
          - добавляемые: позиции из df2, отсутствующие в df1 и со значением > 0
          ✅ Удаляет дубликаты, сохраняя порядок
        """
        if value_col is None:
            value_col = self.value_column[0]

        d1, d2 = self._align()

        # Маски для отсутствия
        absent_in_d2 = d2[self.compare_column].isna() | (d2[self.compare_column].astype(str).str.strip() == "")
        absent_in_d1 = d1[self.compare_column].isna() | (d1[self.compare_column].astype(str).str.strip() == "")

        # Маски для значений > 0
        gt0_d1 = d1[value_col].fillna(0) > 0
        gt0_d2 = d2[value_col].fillna(0) > 0

        # Итоговые списки
        removed = d1.loc[absent_in_d2 & gt0_d1, self.compare_column].astype(str).tolist()
        added = d2.loc[absent_in_d1 & gt0_d2, self.compare_column].astype(str).tolist()

        # ✅ Удаляем дубликаты, сохраняя порядок
        def unique_ordered(seq):
            seen = set()
            return [x for x in seq if not (x in seen or seen.add(x))]

        return unique_ordered(removed), unique_ordered(added)

    def export_added_removed_positions(
            self,
            path: str,
            value_col: Optional[str] = None
    ):
        if value_col is None:
            value_col = self.value_column[0]

        d1, d2 = self._align()

        def unique_ordered(seq):
            seen = set()
            return [x for x in seq if not (x in seen or seen.add(x))]

        unique1 = unique_ordered(
            d1.loc[d1[value_col].fillna(0) > 0, self.compare_column].astype(str).tolist()
        )
        unique2 = unique_ordered(
            d2.loc[d2[value_col].fillna(0) > 0, self.compare_column].astype(str).tolist()
        )

        removed = [x for x in unique1 if x not in unique2]
        added = [x for x in unique2 if x not in unique1]

        def make_df_with_index(values, col_name):
            return pd.DataFrame({
                "№": range(1, len(values) + 1),
                col_name: values
            })

        import pandas as pd
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            make_df_with_index(removed, "Исключаемые позиции").to_excel(
                writer, sheet_name="Исключаемые", index=False
            )
            make_df_with_index(added, "Добавляемые позиции").to_excel(
                writer, sheet_name="Добавляемые", index=False
            )
            make_df_with_index(unique1, "Уникальные df1").to_excel(
                writer, sheet_name="Уникальные df1", index=False
            )
            make_df_with_index(unique2, "Уникальные df2").to_excel(
                writer, sheet_name="Уникальные df2", index=False)
# demo
def compare_smetas(
    project_path: str,
    fact_path: str,
    *,
    compare_column: str = "Наименование",
    value_column: Union[str, List[str]] = "Стоимость",
    extra_column: Optional[List[str]] = ["Единица измерения"],
    subsection_column: Optional[str] = "Подраздел",
) -> pd.DataFrame:
    """
    Быстро обработать ДВЕ сметы, сравнить их и вернуть
    DataFrame «отчёта для заказчика».
    """
    # 1. обработка обоих файлов «как обычно»
    df_project = process_smeta(project_path)
    df_fact    = process_smeta(fact_path)

    # 2. сравнение
    cmp = SmetaComparator(
        df_project,
        df_fact,
        file1_name=os.path.basename(project_path),
        file2_name=os.path.basename(fact_path),
        compare_column=compare_column,
        value_column=value_column,
        extra_column=extra_column,
        subsection_column=subsection_column,
    )
    return cmp.generate_customer_report()
