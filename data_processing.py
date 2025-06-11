"""
data_processing.py
Чистый парсер смет: Smeta.parse() → pandas.DataFrame.
Работает на Python 3.7+ (без использования оператора `|`).
"""

import os
import re
import logging
import warnings
from dataclasses import dataclass
from typing import Optional, Dict, List, Any

import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string

# ──────────────────  НАСТРОЙКИ  ──────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    message="Workbook contains no default style"
)

SECTION_RE = re.compile(r"^\s*раздел\s+\d+\.\s*.+", re.IGNORECASE)

CATEGORY_PREFIXES: Dict[str, List[str]] = {
    "Материалы": ["ТССЦ", "ТЦ", "ФССЦ", "ФСБЦ","Прайс","Прайслист"],
    "Механизмы": ["ТСЭМ", "ФСЭМ"],
    "Работа":    ["ГЭСН", "ТЕР", "ФЕР"],
}

# ──────────────────  1. КОЛОНКИ  ──────────────────
@dataclass
class ColumnConfig:
    pos:  str = 'A'
    code: str = 'B'
    name: str = 'C'
    unit: str = 'F'
    qty:  str = 'G'
    cost: str = 'N'

    def indexes(self) -> Dict[str, int]:
        """Буквы Excel → 0-based индексы DataFrame/openpyxl."""
        return {k: column_index_from_string(v) - 1 for k, v in self.__dict__.items()}


# ──────────────────  2. КЛАСС SMETA  ──────────────────
class Smeta:
    def __init__(self, file_path: str, config: Optional[ColumnConfig] = None):
        if not os.path.exists(file_path):
            raise FileNotFoundError(file_path)

        self.file_path = file_path
        self.cfg: ColumnConfig = config or ColumnConfig()
        self.wb = openpyxl.load_workbook(file_path, data_only=True)
        self.sheet = self.wb.active

        # ① найдём строку-раздел и, при необходимости, поменяем схему колонок
        self.section_start_row: Optional[int] = (
            self._find_section_start_and_maybe_switch_columns()
        )
        self.idx = self.cfg.indexes()
        logging.info("Карта колонок: %s", self.idx)

        self.data = pd.DataFrame(columns=[
            "Раздел", "Подраздел", "Номер позиции", "Код расценки", "Наименование",
            "Категория", "Единица измерения", "Количество", "Стоимость",
            "ФОТ", "ЭМ", "Материалы", "НР", "СП", "ОТм",
            "Вспомогательные ресурсы", "Оборудование"
        ])

    # ───── 2.1 Найти первую строку-раздел ─────
    def _find_section_start_and_maybe_switch_columns(self) -> Optional[int]:
        """Возвратить строку первого 'Раздел N.' и, если нужно, скорректировать cfg."""
        for row in self.sheet.iter_rows(min_row=1, max_col=1):
            val = row[0].value
            if isinstance(val, str) and SECTION_RE.match(val):
                section_row = row[0].row

                # смотрим строку выше — если в H стоит «4», меняем схему
                if section_row > 1:
                    h_above = self.sheet.cell(
                        row=section_row - 1,
                        column=column_index_from_string('H')
                    ).value
                    if str(h_above).strip() == '4':
                        logging.info("Маркер '4' найден — переключаемся на H/I/P.")
                        self.cfg.unit, self.cfg.qty, self.cfg.cost = 'H', 'I', 'P'
                return section_row
        return None

    # ───── 2.2 Детализация затрат ─────
    RE_VTCH_OTM = re.compile(r"^\s*в\s*т\.ч\.?\s*отм", re.IGNORECASE)

    def _process_cost_details(self, row: List[Any]) -> Dict[str, Optional[float]]:
        cost_mapping = {
            "ОТ":  "ФОТ",
            "ЭМ":  "ЭМ",
            "М":   "Материалы",
            "ОТм": "ОТм",
            "НР":  "НР",
            "СП":  "СП",
            "Вспомогательные ненормируемые материальные ресурсы": "Вспомогательные ресурсы",
        }
        details = {v: None for v in cost_mapping.values()}

        name_col, cost_col = self.idx["name"], self.idx["cost"]
        text = str(row[name_col]).strip()
        cost = row[cost_col] or 0

        # 1) «в т.ч. ОТм …»
        if text.lower().startswith("в т.ч.") and "отм" in text.lower():
            details["ОТм"] = cost
            details["ЭМ"]  = -cost
            return details

        # 2) «ОТм(ЗТм) …»
        if text.startswith("ОТм(ЗТм)"):
            details["ОТм"] = cost
            details["ФОТ"] = -cost
            return details

        # 3) обычные ключи
        for prefix, target in cost_mapping.items():
            if text.startswith(prefix):
                details[target] = cost
                break
        return details

    # ───── 2.3 Категория позиции ─────
    def _get_category(self, cell_a: Any, cell_b: Any) -> str:
        if isinstance(cell_a, str) and "\nО" in cell_a:
            return "Оборудование"

        if isinstance(cell_b, str):
            for cat, prefixes in CATEGORY_PREFIXES.items():
                if any(cell_b.startswith(p) for p in prefixes):
                    return cat
        return "Неизвестная категория"

    # ───── 2.4 Основной разбор ─────
    def parse(self) -> pd.DataFrame:
        if not self.section_start_row:
            raise ValueError("Не найден ни один корректный «Раздел N.»")

        rows = self.sheet.iter_rows(min_row=self.section_start_row, values_only=True)

        pos_c, code_c, name_c = self.idx['pos'], self.idx['code'], self.idx['name']
        unit_c, qty_c, cost_c = self.idx['unit'], self.idx['qty'], self.idx['cost']

        current_section = None
        current_subsection = None
        parsing_position = False
        position_data: Dict[str, Any] = {}

        for row in rows:
            cell_a, cell_b, cell_c = row[pos_c], row[code_c], row[name_c]

            # --- новый Раздел ---
            if isinstance(cell_a, str) and SECTION_RE.match(cell_a):
                current_section = cell_a.strip()
                current_subsection = current_section
                continue

            # --- новый Подраздел ---
            if isinstance(cell_a, str) and not cell_b and not cell_c:
                current_subsection = cell_a.strip()
                continue

            # --- позиция ---
            is_position = (
                (isinstance(cell_a, (int, float)) or
                 (isinstance(cell_a, str) and (cell_a.isdigit() or "\nО" in cell_a)))
                and cell_b
            )
            if is_position:
                if parsing_position:
                    self.data = pd.concat([self.data, pd.DataFrame([position_data])])
                parsing_position = True

                category = self._get_category(cell_a, cell_b)
                position_data = {
                    "Раздел": current_section,
                    "Подраздел": current_subsection,
                    "Номер позиции": cell_a,
                    "Код расценки": cell_b,
                    "Наименование": cell_c,
                    "Категория": category,
                    "Единица измерения": row[unit_c],
                    "Количество": row[qty_c],
                    "Стоимость": None,
                    "ФОТ": None,
                    "ЭМ": None,
                    "Материалы": None,
                    "НР": None,
                    "СП": None,
                    "ОТм": None,
                    "Вспомогательные ресурсы": None,
                    "Оборудование": None,
                }
                continue

            # --- «Всего по позиции» ---
            if parsing_position and cell_c == "Всего по позиции":
                position_data["Стоимость"] = row[cost_c]

                if position_data["Категория"] == "Материалы":
                    position_data["Материалы"] = row[cost_c]
                elif position_data["Категория"] == "Оборудование":
                    position_data["Оборудование"] = row[cost_c]

                self.data = pd.concat([self.data, pd.DataFrame([position_data])])
                parsing_position = False
                continue

            # --- детали затрат ---
            if parsing_position:
                extras = self._process_cost_details(row)
                for k, v in extras.items():
                    if v is not None:
                        position_data[k] = (position_data[k] or 0) + v

        return self.data


# ──────────────────  3. ФУНКЦИЯ-ОБЁРТКА ДЛЯ GUI  ──────────────────
def process_smeta(path: str) -> pd.DataFrame:
    """Вызывается из main.py; возвращает DataFrame для GUI."""
    fname = os.path.basename(path)
    try:
        sm = Smeta(path)
        sm.parse()
        df = sm.data


        # лог подтвердит общую стоимость, но без создания файлов
        logging.info("∑ Стоимость по '%s': %.2f", fname, df["Стоимость"].sum())
        return df

    except Exception as exc:
        logging.error("Ошибка при обработке '%s': %s", fname, exc)
        raise