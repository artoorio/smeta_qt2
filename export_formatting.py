import html
import math
import re
import textwrap
from typing import Iterable

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DEFAULT_ROW_HEIGHT = 18
NAME_MIN_WIDTH = 24
NAME_MAX_WIDTH = 80
NAME_TARGET_LINES = 4
MAX_ROW_HEIGHT = 54
HEADER_FILL = PatternFill("solid", fgColor="EAF1F8")
GRID_SIDE = Side(style="thin", color="C9D3E0")
GRID_BORDER = Border(left=GRID_SIDE, right=GRID_SIDE, top=GRID_SIDE, bottom=GRID_SIDE)


def _stringify(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value)


def _iter_lines(value: object) -> Iterable[str]:
    text = _stringify(value).replace("\r\n", "\n").replace("\r", "\n")
    parts = text.split("\n")
    return parts or [""]


def _max_word_length(text: str) -> int:
    words = [len(word) for word in text.split()]
    return max(words, default=0)


def _wrapped_line_count(value: object, width_chars: int) -> int:
    width_chars = max(1, int(width_chars))
    total = 0
    for line in _iter_lines(value):
        normalized = line.strip()
        if not normalized:
            total += 1
            continue
        wrapped = textwrap.wrap(
            normalized,
            width=width_chars,
            break_long_words=True,
            break_on_hyphens=False,
        )
        total += max(1, len(wrapped))
    return max(1, total)


def _looks_numeric_header(header: object) -> bool:
    text = str(header)
    return any(
        token in text
        for token in ("Кол-во", "Количество", "Ст-ть", "Стоимость", "Разница", "Материалы", "ФОТ", "ЭМ", "НР", "СП", "ОТм")
    )


def _looks_quantity_header(header: object) -> bool:
    text = str(header)
    return any(token in text for token in ("Кол-во", "Количество"))


def _looks_code_header(header: object) -> bool:
    text = str(header)
    return text in {"№", "Код расценки", "Ед.изм.", "Единица измерения", "№ позиции"}


def _to_float_if_numeric(value: object):
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)) and not pd.isna(value):
        return float(value)
    if isinstance(value, str):
        normalized = value.replace(" ", "").replace("\xa0", "").replace(",", ".").strip()
        if not normalized:
            return None
        try:
            return float(normalized)
        except ValueError:
            return None
    return None


def _format_html_value(header: object, value: object) -> str:
    numeric_value = _to_float_if_numeric(value)
    if numeric_value is not None and _looks_numeric_header(header):
        if _looks_quantity_header(header):
            text = format(numeric_value, ".15g")
            if "e" in text.lower():
                text = f"{numeric_value}"
            if "." in text:
                integer_part, fractional_part = text.split(".", 1)
                integer_text = f"{int(integer_part):,}".replace(",", " ")
                fractional_part = fractional_part.rstrip("0")
                text = f"{integer_text},{fractional_part}" if fractional_part else integer_text
            else:
                text = f"{int(numeric_value):,}".replace(",", " ")
            return html.escape(text)
        if float(numeric_value).is_integer():
            text = f"{int(numeric_value):,}".replace(",", " ")
        else:
            text = f"{numeric_value:,.2f}".replace(",", " ")
        return html.escape(text)
    return html.escape(_stringify(value)).replace("\n", "<br>")


def _header_html(label: object) -> str:
    return html.escape(str(label)).replace("\n", "<br>")


def _is_divider_row(values: list[object]) -> bool:
    for value in values:
        text = _stringify(value).strip()
        if text.startswith("--"):
            return True
    return False


def suggest_name_width(
    values: Iterable[object],
    target_lines: int = NAME_TARGET_LINES,
    min_width: int = NAME_MIN_WIDTH,
    max_width: int = NAME_MAX_WIDTH,
) -> int:
    max_len = 0
    max_word = 0
    for value in values:
        for line in _iter_lines(value):
            max_len = max(max_len, len(line))
            max_word = max(max_word, _max_word_length(line))

    base_width = math.ceil(max_len / max(1, target_lines)) + 2
    width = max(base_width, max_word + 1, min_width)
    return min(max_width, width)


def suggest_column_width(
    values: Iterable[object],
    min_width: int = 10,
    max_width: int = 40,
    wrap_lines: int = 1,
) -> int:
    max_len = 0
    max_word = 0
    for value in values:
        for line in _iter_lines(value):
            max_len = max(max_len, len(line))
            max_word = max(max_word, _max_word_length(line))

    if wrap_lines <= 1:
        width = max(max_len + 2, min_width)
    else:
        width = max(math.ceil(max_len / wrap_lines) + 2, max_word + 1, min_width)
    return min(max_width, width)


def dataframe_to_html_table(
    df: pd.DataFrame,
    name_column: str = "Наименование",
    target_lines: int = NAME_TARGET_LINES,
    no_wrap_columns: set[str] | None = None,
    table_title: str | None = None,
) -> str:
    headers = list(df.columns)
    if not headers:
        title_html = f"<h2>{html.escape(table_title)}</h2>" if table_title else ""
        return f'{title_html}<div class="table-empty">Нет данных для отображения.</div>'

    no_wrap_columns = no_wrap_columns or set()
    colgroup_parts: list[str] = []
    for header in headers:
        values = [_stringify(header)] + [_stringify(value) for value in df[header].head(300)]
        if header == name_column:
            width = suggest_name_width(df[header].tolist(), target_lines=target_lines)
            style = f"width:{width}ch;min-width:{width}ch;"
        elif str(header) in no_wrap_columns:
            width = suggest_column_width(values, min_width=14, max_width=120, wrap_lines=1)
            style = f"width:{width}ch;min-width:{width}ch;"
        elif _looks_numeric_header(header):
            width = suggest_column_width(values, min_width=10, max_width=20, wrap_lines=2 if '\n' in str(header) else 1)
            style = f"width:{width}ch;min-width:{max(9, width - 2)}ch;"
        elif _looks_code_header(header):
            width = suggest_column_width(values, min_width=10, max_width=22, wrap_lines=1)
            style = f"width:{width}ch;min-width:{width}ch;"
        else:
            width = suggest_column_width(values, min_width=12, max_width=36, wrap_lines=2 if '\n' in str(header) else 1)
            style = f"width:{width}ch;min-width:{max(12, width - 2)}ch;"
        colgroup_parts.append(f'<col style="{style}">')

    thead = "".join(
        f'<th class="{"col-numeric" if _looks_numeric_header(header) else "col-text"}">{_header_html(header)}</th>'
        for header in headers
    )

    body_rows: list[str] = []
    qty_header = "Количество" if "Количество" in headers else None
    cat_header = "Категория" if "Категория" in headers else None
    for _, row in df.iterrows():
        row_values = [row[header] for header in headers]
        row_classes: list[str] = []
        if _is_divider_row(row_values):
            row_classes.append("row-divider")
        elif cat_header:
            category = _stringify(row.get(cat_header)).strip()
            qty_value = _to_float_if_numeric(row.get(qty_header)) if qty_header else None
            if category == "Работа":
                row_classes.append("row-work")
            elif category == "Материалы":
                row_classes.append("row-material-negative" if qty_value is not None and qty_value < 0 else "row-material")
            elif category == "Механизмы":
                row_classes.append("row-machinery")

        cells = []
        for header in headers:
            classes = ["col-numeric" if _looks_numeric_header(header) else "col-text"]
            if header == name_column:
                classes.append("col-name")
            if str(header) in no_wrap_columns:
                classes.append("no-wrap")
            value_html = _format_html_value(header, row[header])
            if str(header).startswith("Разница"):
                numeric_value = _to_float_if_numeric(row[header])
                if numeric_value is not None:
                    classes.append("diff-positive" if numeric_value > 0 else "diff-negative" if numeric_value < 0 else "diff-zero")
            cells.append(f'<td class="{" ".join(classes)}">{value_html}</td>')
        class_attr = f' class="{" ".join(row_classes)}"' if row_classes else ""
        body_rows.append(f"<tr{class_attr}>{''.join(cells)}</tr>")

    title_html = f"<h2>{html.escape(table_title)}</h2>" if table_title else ""
    return (
        f'{title_html}<div class="table-scroll"><table class="processed-table"><colgroup>'
        f'{"".join(colgroup_parts)}</colgroup><thead><tr>{thead}</tr></thead>'
        f'<tbody>{"".join(body_rows)}</tbody></table></div>'
    )


def dataframes_to_readable_html(
    sections: list[tuple[str, pd.DataFrame]],
    title: str = "Отчёт",
    name_column: str = "Наименование",
    target_lines: int = NAME_TARGET_LINES,
    no_wrap_columns: dict[str, set[str]] | None = None,
) -> str:
    no_wrap_columns = no_wrap_columns or {}
    sections_html = []
    for section_title, frame in sections:
        sections_html.append(
            '<section class="report-section">'
            + dataframe_to_html_table(
                frame,
                name_column="Файл" if "Файл" in frame.columns else name_column,
                target_lines=target_lines,
                no_wrap_columns=no_wrap_columns.get(section_title, set()),
                table_title=section_title,
            )
            + "</section>"
        )

    return f"""<!DOCTYPE html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(title)}</title>
  <style>
    :root {{
      --page-bg: #f7f4ee;
      --page-accent: #ddc7aa;
      --surface: #fffdfa;
      --surface-strong: #f3ece1;
      --header: #eaf1f8;
      --grid: #c9d3e0;
      --text: #1f2933;
      --muted: #5c6b78;
      --work: #d6efd6;
      --material: #d6eaf8;
      --material-negative: #f8d6d6;
      --machinery: #fff6d6;
      --divider: #e9ecef;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: "Segoe UI", "Trebuchet MS", Arial, sans-serif;
      color: var(--text);
      background:
        radial-gradient(circle at top left, rgba(221, 199, 170, 0.42), transparent 32%),
        linear-gradient(180deg, #fbf8f3 0%, var(--page-bg) 100%);
    }}
    .page {{
      max-width: 1520px;
      margin: 0 auto;
      padding: 28px 24px 40px;
    }}
    .page-header {{
      margin-bottom: 20px;
      padding: 18px 22px;
      border: 1px solid rgba(201, 211, 224, 0.9);
      border-radius: 22px;
      background: linear-gradient(180deg, rgba(255, 253, 250, 0.98), rgba(247, 244, 238, 0.96));
      box-shadow: 0 12px 30px rgba(95, 83, 64, 0.08);
    }}
    h1 {{
      margin: 0;
      font-size: 28px;
      line-height: 1.2;
    }}
    .subtitle {{
      margin: 8px 0 0;
      color: var(--muted);
      font-size: 14px;
    }}
    .report-section {{
      margin-top: 18px;
      padding: 16px;
      border: 1px solid rgba(201, 211, 224, 0.9);
      border-radius: 22px;
      background: rgba(255, 253, 250, 0.96);
      box-shadow: 0 10px 24px rgba(95, 83, 64, 0.06);
    }}
    h2 {{
      margin: 0 0 12px;
      font-size: 18px;
    }}
    .table-scroll {{
      overflow-x: auto;
      border: 1px solid var(--grid);
      border-radius: 18px;
      background: white;
    }}
    .processed-table {{
      width: 100%;
      border-collapse: collapse;
      table-layout: fixed;
      font-size: 14px;
    }}
    .processed-table th,
    .processed-table td {{
      padding: 8px 10px;
      border: 1px solid var(--grid);
      vertical-align: middle;
    }}
    .processed-table th {{
      background: var(--header);
      position: sticky;
      top: 0;
      z-index: 1;
      font-weight: 700;
      white-space: normal;
    }}
    .processed-table td {{
      white-space: normal;
      word-break: break-word;
    }}
    .processed-table .col-text {{ text-align: left; }}
    .processed-table .col-numeric {{ text-align: right; }}
    .processed-table .col-name {{ white-space: normal; }}
    .processed-table .no-wrap {{ white-space: nowrap; word-break: normal; }}
    .processed-table tr.row-divider td {{
      background: var(--divider);
      font-weight: 700;
    }}
    .processed-table tr.row-work td {{ background: var(--work); }}
    .processed-table tr.row-material td {{ background: var(--material); }}
    .processed-table tr.row-material-negative td {{ background: var(--material-negative); }}
    .processed-table tr.row-machinery td {{ background: var(--machinery); }}
    .processed-table td.diff-positive {{ color: #1e7a42; font-weight: 600; }}
    .processed-table td.diff-negative {{ color: #b42318; font-weight: 600; }}
    .processed-table td.diff-zero {{ color: var(--muted); }}
    .table-empty {{
      padding: 16px 18px;
      border: 1px dashed var(--grid);
      border-radius: 16px;
      color: var(--muted);
      background: rgba(243, 236, 225, 0.45);
    }}
    @media (max-width: 720px) {{
      .page {{ padding: 16px 12px 28px; }}
      .report-section {{ padding: 12px; }}
      .processed-table {{ font-size: 13px; }}
    }}
  </style>
</head>
<body>
  <main class="page">
    <header class="page-header">
      <h1>{html.escape(title)}</h1>
      <p class="subtitle">HTML-отчёт оформлен по той же логике, что и Excel: читаемая шапка, переносы текста, границы и выравнивание чисел.</p>
    </header>
    {''.join(sections_html)}
  </main>
</body>
</html>
"""


def apply_readable_sheet_layout(
    ws,
    df: pd.DataFrame,
    name_column: str = "Наименование",
    target_lines: int = NAME_TARGET_LINES,
    no_wrap_columns: set[str] | None = None,
    max_row_height: float | None = MAX_ROW_HEIGHT,
) -> None:
    headers = list(df.columns)
    if not headers:
        return
    no_wrap_columns = no_wrap_columns or set()

    name_index = headers.index(name_column) + 1 if name_column in headers else None
    name_width = (
        suggest_name_width(df[name_column].tolist(), target_lines=target_lines)
        if name_index
        else None
    )

    for col_index, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_index)
        if col_index == name_index and name_width is not None:
            ws.column_dimensions[column_letter].width = name_width
            continue

        values = [_stringify(header)] + [_stringify(value) for value in df[header].head(300)]
        if str(header) in no_wrap_columns:
            width = suggest_column_width(values, min_width=14, max_width=80, wrap_lines=1)
        else:
            width = suggest_column_width(values, min_width=10, max_width=40, wrap_lines=2 if "\n" in str(header) else 1)
        ws.column_dimensions[column_letter].width = width

    for header_cell in ws[1]:
        header_cell.alignment = Alignment(wrapText=True, vertical="center")
        header_cell.font = Font(bold=True)
        header_cell.fill = HEADER_FILL
        header_cell.border = GRID_BORDER

    ws.freeze_panes = "A2"

    if not name_index or name_width is None:
        return

    for row_index in range(2, len(df) + 2):
        cell = ws.cell(row=row_index, column=name_index)
        cell.alignment = Alignment(wrapText=True, vertical="center")
        line_count = _wrapped_line_count(cell.value, name_width)
        target_height = max(DEFAULT_ROW_HEIGHT, line_count * 15)
        if max_row_height is not None:
            target_height = min(max_row_height, target_height)
        ws.row_dimensions[row_index].height = target_height

    for row in ws.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = GRID_BORDER
            column_name = headers[cell.column - 1]
            if cell.column != name_index:
                if str(column_name) in no_wrap_columns:
                    cell.alignment = Alignment(vertical="center", wrapText=False)
                else:
                    cell.alignment = Alignment(vertical="center")


def apply_named_column_widths(ws, headers: list[str], width_map: dict[str, int | float]) -> None:
    for col_index, header in enumerate(headers, start=1):
        width = width_map.get(str(header))
        if width is not None:
            ws.column_dimensions[get_column_letter(col_index)].width = width


def autofit_columns_by_name(
    ws,
    df: pd.DataFrame,
    columns: list[str],
    min_width: int = 10,
    max_width: int = 80,
    wrap_lines: int = 1,
) -> None:
    headers = list(df.columns)
    for column in columns:
        if column not in headers:
            continue
        values = [column] + [_stringify(value) for value in df[column].head(300)]
        width = suggest_column_width(
            values,
            min_width=min_width,
            max_width=max_width,
            wrap_lines=wrap_lines,
        )
        ws.column_dimensions[get_column_letter(headers.index(column) + 1)].width = width


def cap_row_heights(ws, max_height: float) -> None:
    for row_index, dimension in ws.row_dimensions.items():
        if dimension.height and dimension.height > max_height:
            dimension.height = max_height


def apply_section_row_grouping(ws, df: pd.DataFrame, divider_column: str) -> None:
    if divider_column not in df.columns:
        return

    divider_values = df[divider_column].fillna("").astype(str).tolist()
    section_rows: list[int] = []
    for idx, value in enumerate(divider_values, start=2):
        text = value.strip()
        if text.startswith("-- ") and not text.startswith("---- "):
            section_rows.append(idx)

    if not section_rows:
        return

    ws.sheet_properties.outlinePr.summaryBelow = False

    for i, section_row in enumerate(section_rows):
        start_row = section_row + 1
        end_row = (section_rows[i + 1] - 1) if i + 1 < len(section_rows) else (len(df) + 1)
        if start_row <= end_row:
            ws.row_dimensions.group(start_row, end_row, outline_level=1, hidden=False)


def dataframe_to_readable_html(
    df: pd.DataFrame,
    title: str = "Обработанная смета",
    name_column: str = "Наименование",
    target_lines: int = NAME_TARGET_LINES,
) -> str:
    return dataframes_to_readable_html(
        [("Данные", df)],
        title=title,
        name_column=name_column,
        target_lines=target_lines,
    )
